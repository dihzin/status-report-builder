import openpyxl

REQUIRED_SHEETS = {
    "CONFIG": ["project_name", "sponsor", "report_date", "owner_name"],
    "FASES": ["ordem", "nome", "status", "data_alvo", "destaque"],
    "KPIS": ["ordem", "titulo", "valor", "subtitulo", "tipo", "nivel"],
    "RESUMO_EXECUTIVO": ["ordem", "texto", "status"],
    "PENDENCIAS_CRITICAS": ["prioridade", "item", "responsaveis", "status", "nivel"],
    "PROXIMAS_ACOES": ["ordem", "texto"],
    "CURVA_S": ["dia", "planejado", "realizado"],
    "MARCOS": ["ordem", "nome", "data_alvo", "status", "tipo"],
    "RODAPE": [],
}

OPTIONAL_SHEETS = {
    "BRANDING": ["cor_primaria"],
    "PRESENTATION_CONFIG": [],
    "GANTT_TAREFAS": ["id", "nome", "inicio", "fim"],
    "GANTT_MARCOS": ["id", "nome", "data"],
    "GANTT_CONFIG": [],
}


# Abas que usam formato chave-valor (coluna A = chave, coluna B = valor)
_KV_SHEETS = {"CONFIG", "RODAPE"}


def _headers(ws):
    cells = [c for c in ws[1] if c.value is not None]
    return [str(c.value).strip() for c in cells]


def _kv_keys(ws):
    """Retorna os campos (coluna A) de uma aba no formato chave-valor."""
    return [
        str(row[0]).strip()
        for row in ws.iter_rows(min_row=1, max_col=1, values_only=True)
        if row[0] is not None
    ]


def validate_schema(filepath):
    errors = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)

        for sheet_name, required_cols in REQUIRED_SHEETS.items():
            if sheet_name not in wb.sheetnames:
                errors.append(f"Aba obrigatória ausente: {sheet_name}")
                continue
            if required_cols:
                if sheet_name in _KV_SHEETS:
                    # Aba chave-valor: verifica presença das chaves na coluna A
                    keys = _kv_keys(wb[sheet_name])
                    for col in required_cols:
                        if col not in keys:
                            errors.append(f"Aba '{sheet_name}': campo obrigatório '{col}' não encontrado")
                else:
                    headers = _headers(wb[sheet_name])
                    for col in required_cols:
                        if col not in headers:
                            errors.append(f"Aba '{sheet_name}': coluna obrigatória '{col}' não encontrada")

        _KV_OPTIONAL = {"BRANDING"}
        for sheet_name, expected_cols in OPTIONAL_SHEETS.items():
            if sheet_name not in wb.sheetnames:
                continue
            if sheet_name in _KV_OPTIONAL:
                keys = _kv_keys(wb[sheet_name])
                for col in expected_cols:
                    if col not in keys:
                        errors.append(f"Aba opcional '{sheet_name}': campo esperado '{col}' não encontrado")
            else:
                headers = _headers(wb[sheet_name])
                for col in expected_cols:
                    if col not in headers:
                        errors.append(f"Aba opcional '{sheet_name}': coluna esperada '{col}' não encontrada")

        wb.close()
    except PermissionError:
        errors.append("Arquivo Excel bloqueado — validação ignorada (usando cache).")
    except Exception as e:
        errors.append(f"Erro ao validar schema: {str(e)}")

    return errors
