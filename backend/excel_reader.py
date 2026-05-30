import os
from datetime import date
import openpyxl

_cache = None


def _nav_row(row):
    """Retorna True se a linha for o link de navegação da capa (← Voltar à Capa)."""
    v = str(row[0] or '').strip()
    return 'Voltar' in v or v.startswith('←')


def _read_key_value_sheet(ws):
    data = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] is None or len(row) < 2:
            continue
        # Pula linha de navegação inserida pelo create_capa.py
        if _nav_row(row):
            continue
        key = str(row[0]).strip()
        # Pula linha de cabeçalho visual e linhas de legenda (col B vazia/mesclada)
        if not key or key.lower() == 'campo' or row[1] is None:
            continue
        data[key] = row[1]
    return data


def _read_table_sheet(ws, require_col0=True):
    # Detecta se row 1 é link de navegação; se sim, usa row 2 como cabeçalho
    first_row_vals = [c.value for c in ws[1]]
    if first_row_vals and _nav_row(first_row_vals):
        header_row_idx = 2
        data_start_row = 3
    else:
        header_row_idx = 1
        data_start_row = 2

    headers_row = [c for c in ws[header_row_idx] if c.value is not None]
    headers = [str(c.value).strip() if c.value else f"_col{i}" for i, c in enumerate(headers_row)]
    rows = []
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        # require_col0=False: aceita linhas onde col 0 é None, mas alguma outra coluna tem dado
        if require_col0:
            if row[0] is None:
                continue
        else:
            if not any(v is not None for v in row[:len(headers)]):
                continue
        item = {}
        for i, h in enumerate(headers):
            item[h] = row[i] if i < len(row) else None
        rows.append(item)
    return rows


def read_excel(filepath):
    global _cache

    if not os.path.exists(filepath):
        return None, "Arquivo Excel não encontrado"

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except PermissionError:
        if _cache is not None:
            return _cache, "Arquivo bloqueado (aberto no Excel). Usando último cache válido."
        return None, "Arquivo Excel está bloqueado (aberto no Excel). Feche o arquivo e tente novamente."
    except Exception as e:
        if _cache is not None:
            return _cache, f"Erro ao ler Excel: {str(e)}. Usando último cache válido."
        return None, f"Erro ao ler Excel: {str(e)}"

    try:
        data = {}

        data["branding"] = _read_key_value_sheet(wb["BRANDING"]) if "BRANDING" in wb.sheetnames else {}

        if "CONFIG" in wb.sheetnames:
            data["config"] = _read_key_value_sheet(wb["CONFIG"])
        else:
            data["config"] = {}

        if "PRESENTATION_CONFIG" in wb.sheetnames:
            data["presentation_config"] = _read_key_value_sheet(wb["PRESENTATION_CONFIG"])
        else:
            data["presentation_config"] = {}

        sheet_map = {
            "FASES":                "fases",
            "KPIS":                 "kpis",
            "RESUMO_EXECUTIVO":     "resumo_executivo",
            "PENDENCIAS_CRITICAS":  "pendencias_criticas",
            "PROXIMAS_ACOES":       "proximas_acoes",
            "MARCOS":               "marcos",
            "GANTT_TAREFAS":        "gantt_tarefas",
            "GANTT_MARCOS":         "gantt_marcos",
        }

        for sheet_name, key in sheet_map.items():
            if sheet_name in wb.sheetnames:
                data[key] = _read_table_sheet(wb[sheet_name])
            else:
                data[key] = []

        # CURVA_S: lê linhas onde 'dia' pode ser None (usuário preenche só na 1ª)
        if "CURVA_S" in wb.sheetnames:
            cs_raw = _read_table_sheet(wb["CURVA_S"], require_col0=False)
            # Auto-gera valores de 'dia' para linhas onde está ausente.
            # Distribui de forma linear entre o 1º e último dia do projeto.
            total_days = int(data.get("config", {}).get("total_days") or 0) or 64
            n = len(cs_raw)
            if n > 0:
                # Se o usuário preencheu todos os dias, mantém; se não, distribui.
                filled = [r for r in cs_raw if r.get("dia") is not None]
                if len(filled) == 0:
                    # Nenhum dia preenchido: distribui de 1 a total_days
                    for i, r in enumerate(cs_raw):
                        r["dia"] = round(1 + i * (total_days - 1) / max(n - 1, 1))
                elif len(filled) < n:
                    # Alguns dias preenchidos: interpola os restantes
                    last_dia = 1
                    for i, r in enumerate(cs_raw):
                        if r.get("dia") is not None:
                            last_dia = int(r["dia"])
                        else:
                            # Estima próximo ponto proporcional
                            remaining = n - i
                            last_dia = last_dia + max(1, round((total_days - last_dia) / max(remaining, 1)))
                            r["dia"] = last_dia
            data["curva_s"] = cs_raw
        else:
            data["curva_s"] = []

        if "RODAPE" in wb.sheetnames:
            data["rodape"] = _read_key_value_sheet(wb["RODAPE"])
        else:
            data["rodape"] = {}

        if "GANTT_CONFIG" in wb.sheetnames:
            data["gantt_config"] = _read_key_value_sheet(wb["GANTT_CONFIG"])
        else:
            data["gantt_config"] = {}

        wb.close()
        _cache = data
        return data, None

    except Exception as e:
        wb.close()
        if _cache is not None:
            return _cache, f"Erro ao processar dados: {str(e)}. Usando último cache válido."
        return None, f"Erro ao processar dados: {str(e)}"


def create_template(filepath):
    today = date.today().strftime("%d/%m/%Y")

    wb = openpyxl.Workbook()

    # ── BRANDING ──────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "BRANDING"
    for key, val in [
        ("cor_primaria",   "#2a7249"),
        ("cor_secundaria", "#1a4f35"),
        ("cor_texto",      "#ffffff"),
        ("logo_path",      "assets/logo.svg"),
    ]:
        ws.append([key, val])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 36

    # ── CONFIG ────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("CONFIG")
    config_data = [
        ("logo_path",        "assets/logo.svg"),
        ("report_title",     "Status Report"),
        ("project_name",     "Projeto Exemplo"),
        ("project_subtitle", "Transformação Digital"),
        ("sponsor",          "Sponsor do Projeto"),
        ("pmar_source",      "docs/PMAR.45_RAID_Log.xlsm"),
        ("last_pmar_import", ""),
        ("alert_label",      "Atenção"),
        ("alert_level",      "warning"),
        ("report_date",      today),
        ("current_phase",    "Explorer"),
        ("current_day",      7),
        ("total_days",       46),
        ("progress_percent", 42),
        ("owner_name",       "Nome do Responsável"),
        ("report_name",      "Status Report – Fase Explorer"),
        ("partner_name",     "Stratesys"),
        ("presentation_duration", "30 minutos"),
        ("cover_eyebrow",    "STATUS REPORT DO PROJETO · FASE EXPLORE"),
        ("cover_main_title", "Implementação|SAP Ariba na EDF"),
        ("cover_subtitle",   "Procurement digital integrado ao S/4HANA."),
        ("cover_highlight",  "EDF"),
        ("cover_tagline",    "EDF × Stratesys, parceria estratégica"),
        ("cover_restriction_label", "USO RESTRITO"),
        ("cover_footer_left", "Implementação SAP Ariba"),
        ("cover_footer_right", "EDF × Stratesys"),
    ]
    for key, val in config_data:
        ws.append([key, val])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40

    # ── FASES ─────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("FASES")
    ws.append(["ordem", "nome", "status", "data_inicio", "data_alvo", "destaque"])
    fases = [
        [1, "Prepare",   "Concluído",    "25/03/2026", "01/04/2026", False],
        [2, "Explorer",  "Em andamento", "02/04/2026", today,        True],
        [3, "Realize",   "Planejado",    "15/05/2026", "05/06/2026", False],
        [4, "SIT",       "Planejado",    "06/06/2026", "20/06/2026", False],
        [5, "UAT",       "Planejado",    "21/06/2026", "05/07/2026", False],
        [6, "Deploy",    "Planejado",    "06/07/2026", "20/07/2026", False],
        [7, "Hypercare", "Planejado",    "21/07/2026", "31/07/2026", False],
    ]
    for row in fases:
        ws.append(row)

    # ── KPIS ──────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("KPIS")
    ws.append(["ordem", "titulo", "valor", "subtitulo", "tipo", "nivel"])
    kpis = [
        [1, "Data",        today,          "",  "calendar", "success"],
        [2, "Fase Atual",  "Explorer",     "",  "compass",  "success"],
        [3, "Progresso",   "42%",          "",  "progress", "success"],
        [4, "Prazo",       "Dia 7 de 46",  "",  "flag",     "success"],
        [5, "Risco Atual", "2 em atenção", "",  "warning",  "warning"],
        [6, "Saúde Geral", "Atenção",      "",  "heart",    "warning"],
    ]
    for row in kpis:
        ws.append(row)

    # ── RESUMO_EXECUTIVO ──────────────────────────────────────────────────────
    ws = wb.create_sheet("RESUMO_EXECUTIVO")
    ws.append(["ordem", "texto", "status"])
    resumo = [
        [1, "Perfis SAP liberados para consultores",           "concluido"],
        [2, "Numeração PARA das novas contas definida",        "concluido"],
        [3, "Estrutura do PDD avançada nas frentes 1.1 e 1.2", "concluido"],
        [4, "Gate Explorer → Realize em preparação",           "andamento"],
    ]
    for row in resumo:
        ws.append(row)

    # ── PENDENCIAS_CRITICAS ───────────────────────────────────────────────────
    # Inclui data_limite (campo data_limite = DUE DATE da PMAR)
    ws = wb.create_sheet("PENDENCIAS_CRITICAS")
    ws.append([
        "prioridade", "item", "responsaveis", "status", "nivel",
        "id_origem", "categoria", "score", "probabilidade", "impacto",
        "estrategia", "data_limite", "comentarios",
    ])
    pendencias = [
        ["P1", "Impactos financeiros e patrimoniais incompletos", "Rodrigo / Bruna",
         "Atrasado",  "danger",  None, None, None, None, None, None, None, None],
        ["P2", "Análise de sistemas legados incompleta",          "Marcelo / Adnan",
         "Em atenção", "warning", None, None, None, None, None, None, None, None],
        ["P3", "Parametrização inicial das novas contas",         "Time Contábil / TI",
         "No prazo",  "success", None, None, None, None, None, None, None, None],
    ]
    for row in pendencias:
        ws.append(row)

    # ── PROXIMAS_ACOES ────────────────────────────────────────────────────────
    ws = wb.create_sheet("PROXIMAS_ACOES")
    ws.append(["ordem", "texto"])
    acoes = [
        [1, "Receber impactos financeiros e patrimoniais completos"],
        [2, "Complementar análise de sistemas legados"],
        [3, "Avançar parametrização inicial das novas contas contábeis"],
        [4, "Consolidar critérios de saída do Explorer"],
    ]
    for row in acoes:
        ws.append(row)

    # ── CURVA_S ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("CURVA_S")
    ws.append(["dia", "planejado", "realizado"])
    # Todos os pontos dentro de 0–100 %
    curva = [
        [1,  0,   0],
        [7,  18,  14],
        [14, 42,  42],
        [21, 60,  50],
        [28, 75,  65],
        [35, 90,  78],
        [42, 100, 92],
        [46, 100, 100],
    ]
    for row in curva:
        ws.append(row)

    # ── MARCOS ────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("MARCOS")
    ws.append(["ordem", "nome", "data_alvo", "status", "tipo"])
    marcos = [
        [1, "Kick-off concluído",      "01/04/2026", "Concluído",    "check"],
        [2, "Gate Explorer → Realize", today,        "Em andamento", "rocket"],
        [3, "Início do SIT",           "20/06/2026", "Planejado",    "gear"],
        [4, "Go-Live",                 "20/07/2026", "Planejado",    "rocket"],
    ]
    for row in marcos:
        ws.append(row)

    # ── PRESENTATION_CONFIG ─────────────────────────────────────────────────────
    ws = wb.create_sheet("PRESENTATION_CONFIG")
    presentation_data = [
        ("font_family",               "Inter, system-ui, -apple-system, sans-serif"),
        ("cover_hero_font_size",      "88"),
        ("cover_hero_line_height",    "1.02"),
        ("cover_subtitle_font_size",  "24"),
        ("cover_eyebrow_font_size",   "14"),
        ("cover_meta_value_size",     "22"),
        ("cover_meta_label_size",     "11"),
        ("alert_warning_bg",          "#E8C86A"),
        ("alert_warning_font",        "#39420A"),
        ("alert_danger_bg",           "#C94A4A"),
        ("alert_success_bg",          "#4A9A63"),
        ("slide_simple_bg",           "#F8FBFD"),
        ("chart_planned_color",       "#3B5F85"),
        ("text_on_dark_primary",      "#FFFFFF"),
        ("text_on_dark_secondary",    "rgba(255,255,255,0.72)"),
        ("text_on_light_primary",     "#1E2228"),
        ("text_on_light_secondary",   "#454B54"),
        ("font_size_alert",           "13"),
        ("font_size_footer",          "12"),
    ]
    for key, val in presentation_data:
        ws.append([key, val])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 56

    # ── RODAPE ────────────────────────────────────────────────────────────────
    # Apenas campos específicos do milestone; owner/nome/data vêm do CONFIG
    ws = wb.create_sheet("RODAPE")
    rodape_data = [
        ("milestone_alvo",  "Explorer → Realize"),
        ("data_alvo",       today),
        ("go_live_previsto", "20/07/2026"),
    ]
    for key, val in rodape_data:
        ws.append([key, val])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40

    # ── GANTT_TAREFAS ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("GANTT_TAREFAS")
    ws.append(["id", "parent_id", "nome", "inicio", "fim", "progresso", "status", "owner", "dependencias"])
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 18

    # ── GANTT_MARCOS ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("GANTT_MARCOS")
    ws.append(["id", "nome", "data", "status", "tipo"])
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12

    # ── GANTT_CONFIG ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("GANTT_CONFIG")
    gantt_cfg_data = [
        ("escala_tempo", "semanas"),
        ("data_inicio_janela", ""),
        ("data_fim_janela", ""),
        ("exibir_baseline", "TRUE"),
        ("exibir_progresso", "TRUE"),
        ("exibir_hoje", "TRUE"),
        ("exibir_dependencias", "FALSE"),
        ("altura_linha", "40"),
        ("largura_dia", "18"),
    ]
    for key, val in gantt_cfg_data:
        ws.append([key, val])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40

    wb.save(filepath)
