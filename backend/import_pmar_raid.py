"""Importa dados de um RAID Log PMAR para o modelo status_projeto.xlsx.

Uso:
    python -m backend.import_pmar_raid
    python -m backend.import_pmar_raid --pmar docs/PMAR.45_RAID_Log.xlsm --status status_projeto.xlsx

A rotina preserva a estrutura do OnePage e adiciona dados RAID em abas opcionais.
"""

from __future__ import annotations

import argparse
import shutil
import warnings
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", message="Data Validation extension is not supported.*")

ROOT_DIR = Path(__file__).resolve().parent.parent
DEFAULT_PMAR_PATH = ROOT_DIR / "docs" / "PMAR.45_RAID_Log.xlsm"
DEFAULT_STATUS_PATH = ROOT_DIR / "status_projeto.xlsx"

ERROR_STRINGS = {"#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"}

PRIORITY_MAP = {
    "critical": ("P1", "danger"),
    "high": ("P2", "warning"),
    "medium": ("P3", "warning"),
    "low": ("P4", "success"),
}

STATUS_MAP = {
    "not started": "Planejado",
    "in progress": "Em atenção",
    "resolved": "Concluído",
    "on hold": "Em atenção",
    "cancelled": "Cancelado",
    "pending": "Pendente",
    "approved": "Aprovado",
    "implemented": "Implementado",
    "rejected": "Rejeitado",
    "deferred": "Adiado",
}

# Cabeçalhos do log de histórico acumulativo (unificação do HISTORICO_IMPORTACAO)
_HIST_HEADERS = [
    "data_importacao", "arquivo_origem", "riscos_importados",
    "riscos_abertos", "riscos_resolvidos", "acoes", "issues", "decisoes",
]
_MAX_HIST = 20


# ---------------------------------------------------------------------------
# Utilitários de limpeza
# ---------------------------------------------------------------------------

def is_error_like(value: Any) -> bool:
    return isinstance(value, str) and value.strip() in ERROR_STRINGS


def clean(value: Any) -> Any:
    if value is None or is_error_like(value):
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def fmt_date(value: Any) -> str | None:
    value = clean(value)
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value)


def date_value(value: Any) -> date | None:
    value = clean(value)
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def to_int(value: Any, default: int = 0) -> int:
    value = clean(value)
    try:
        return int(value)
    except Exception:
        return default


def pluralize(n: int, singular: str, plural: str) -> str:
    return singular if n == 1 else plural


def translate_status(value: Any) -> str:
    raw = str(clean(value) or "").strip()
    return STATUS_MAP.get(raw.lower(), raw)


def map_priority(value: Any) -> tuple[str, str]:
    raw = str(clean(value) or "").strip().lower()
    return PRIORITY_MAP.get(raw, (str(clean(value) or ""), "gray"))


# ---------------------------------------------------------------------------
# Leitura estruturada da PMAR
# ---------------------------------------------------------------------------

def find_header_row(ws, required_headers: Iterable[str]) -> tuple[int, dict[str, int]]:
    required = {h.upper() for h in required_headers}
    for row in ws.iter_rows():
        headers: dict[str, int] = {}
        for cell in row:
            val = clean(cell.value)
            if isinstance(val, str):
                headers[val.upper()] = cell.column
        if required.issubset(set(headers)):
            return row[0].row, headers
    raise ValueError(f"Cabeçalho não encontrado na aba '{ws.title}': {sorted(required)}")


def read_table(ws, required_headers: list[str], description_header: str) -> list[dict[str, Any]]:
    header_row, header_cols = find_header_row(ws, required_headers)
    rows: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        row: dict[str, Any] = {}
        for header, col in header_cols.items():
            row[header] = clean(ws.cell(r, col).value)
        if clean(row.get(description_header.upper())):
            rows.append(row)
    return rows


def read_config(wb) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    ws = wb["Config"]
    config: dict[str, Any] = {}

    # Percorre até 50 linhas para capturar todos os parâmetros
    for r in range(1, min(ws.max_row, 50) + 1):
        key = clean(ws.cell(r, 2).value)
        if isinstance(key, str) and key in {
            "Project Name", "Project Manager", "Sponsor",
            "Start Date", "End Date", "Critical Threshold %",
        }:
            config[key] = clean(ws.cell(r, 3).value)

    # Lê equipe: linhas após o cabeçalho "Name / Role"
    team: list[dict[str, Any]] = []
    team_header_row = None
    for r in range(1, ws.max_row + 1):
        if clean(ws.cell(r, 2).value) == "Name" and clean(ws.cell(r, 3).value) == "Role":
            team_header_row = r
            break
    if team_header_row:
        for r in range(team_header_row + 1, ws.max_row + 1):
            name = clean(ws.cell(r, 2).value)
            role = clean(ws.cell(r, 3).value)
            email = clean(ws.cell(r, 4).value)
            if not name and not role and not email:
                continue
            # Linha de valores de prioridade encerra o bloco de equipe
            if isinstance(name, str) and name in {"Low", "Medium", "High", "Critical"}:
                break
            if name and role:
                team.append({"nome": name, "papel": role, "email": email})
    return config, team


def read_pmar(pmar_path: Path) -> dict[str, Any]:
    if not pmar_path.exists():
        raise FileNotFoundError(f"Arquivo PMAR não encontrado: {pmar_path}")

    required_tabs = ["Config", "Risks", "Actions", "Issues", "Decisions"]
    wb = load_workbook(pmar_path, data_only=True, read_only=False, keep_vba=True)

    missing = [t for t in required_tabs if t not in wb.sheetnames]
    if missing:
        wb.close()
        raise ValueError(f"Abas obrigatórias ausentes na PMAR: {missing}")

    try:
        config, team = read_config(wb)

        risks = read_table(
            wb["Risks"],
            ["ID", "RISK DESCRIPTION", "PRIORITY", "STATUS"],
            "RISK DESCRIPTION",
        )
        actions = read_table(
            wb["Actions"],
            ["ID", "ACTION DESCRIPTION", "STATUS"],
            "ACTION DESCRIPTION",
        )
        issues = read_table(
            wb["Issues"],
            ["ID", "ISSUE DESCRIPTION", "STATUS"],
            "ISSUE DESCRIPTION",
        )
        decisions = read_table(
            wb["Decisions"],
            ["ID", "DECISION", "STATUS"],
            "DECISION",
        )
    finally:
        wb.close()

    return {
        "config": config,
        "team": team,
        "risks": risks,
        "actions": actions,
        "issues": issues,
        "decisions": decisions,
    }


# ---------------------------------------------------------------------------
# Helpers de escrita no workbook
# ---------------------------------------------------------------------------

def ensure_sheet(wb: Workbook, name: str):
    return wb[name] if name in wb.sheetnames else wb.create_sheet(name)


def clear_sheet(ws):
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)


def clear_table_rows(ws, start_row: int = 2):
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def set_kv(ws, key: str, value: Any):
    for row in ws.iter_rows(min_row=1, max_col=2):
        if row[0].value == key:
            row[1].value = value
            return
    ws.append([key, value])


def append_rows(ws, rows: list[list[Any]]):
    for row in rows:
        ws.append(row)


def style_sheet(ws, width_map: dict[str, int] | None = None):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if ws.max_row >= 1:
        for cell in ws[1]:
            if cell.value is not None:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    ws.freeze_panes = "A2"
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = (width_map or {}).get(letter)
        if not width:
            max_len = 0
            for cell in ws[letter]:
                max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            width = min(max(max_len + 2, 10), 42)
        ws.column_dimensions[letter].width = width


def _derive_current_phase(wb: Workbook) -> str | None:
    """Retorna o nome da primeira fase 'Em andamento' da aba FASES, ou None."""
    if "FASES" not in wb.sheetnames:
        return None
    ws = wb["FASES"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome = row[1] if len(row) > 1 else None
        status = row[2] if len(row) > 2 else None
        if nome and status:
            s = str(status).lower()
            if "andamento" in s or "active" in s:
                return str(nome)
    return None


def _backup_excel(path: Path) -> None:
    """Cria backup do Excel em <nome>_backup.xlsx (sobrescreve backup anterior)."""
    backup = path.with_name(path.stem + "_backup.xlsx")
    shutil.copy2(path, backup)


def _update_historico(ws, entry: dict) -> None:
    """Prepend nova entrada no log de importações, mantendo no máximo _MAX_HIST linhas."""
    existing: list[list] = []
    if ws.max_row >= 2:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                existing.append(list(row[: len(_HIST_HEADERS)]))

    new_row = [entry.get(h) for h in _HIST_HEADERS]
    all_rows = ([new_row] + existing)[: _MAX_HIST]

    clear_sheet(ws)
    ws.append(_HIST_HEADERS)
    for row in all_rows:
        ws.append(row)


# ---------------------------------------------------------------------------
# Atualização principal do status_projeto.xlsx
# ---------------------------------------------------------------------------

def update_status_workbook(
    status_path: Path,
    pmar_data: dict[str, Any],
    source_path: Path,
    keep_existing_progress: bool = True,
) -> dict[str, Any]:
    if not status_path.exists():
        raise FileNotFoundError(f"Modelo não encontrado: {status_path}")

    # Backup antes de qualquer escrita
    _backup_excel(status_path)

    wb = load_workbook(status_path)
    config = pmar_data["config"]
    risks = pmar_data["risks"]
    team = pmar_data["team"]

    start_dt = date_value(config.get("Start Date"))
    end_dt = date_value(config.get("End Date"))
    report_dt = date.today()
    current_day = (report_dt - start_dt).days if start_dt else None
    total_days = (end_dt - start_dt).days if start_dt and end_dt else None

    open_risks = [
        r for r in risks
        if str(clean(r.get("STATUS")) or "").lower() not in {"resolved", "cancelled"}
    ]
    resolved_risks = [
        r for r in risks
        if str(clean(r.get("STATUS")) or "").lower() == "resolved"
    ]
    critical_count = sum(1 for r in risks if str(clean(r.get("PRIORITY")) or "").lower() == "critical")
    high_count = sum(1 for r in risks if str(clean(r.get("PRIORITY")) or "").lower() == "high")

    # ── CONFIG ──────────────────────────────────────────────────────────────
    ws_cfg = ensure_sheet(wb, "CONFIG")
    set_kv(ws_cfg, "project_name", config.get("Project Name"))
    # project_subtitle NÃO recebe Sponsor — campo independente, preservado do template
    set_kv(ws_cfg, "sponsor", config.get("Sponsor"))
    set_kv(ws_cfg, "owner_name", config.get("Project Manager"))
    set_kv(ws_cfg, "report_title", "Status Report")
    set_kv(ws_cfg, "alert_label", "ATENÇÃO: riscos PMAR em acompanhamento")
    set_kv(ws_cfg, "alert_level", "danger" if critical_count else "warning")
    set_kv(ws_cfg, "report_date", report_dt.strftime("%d/%m/%Y"))
    set_kv(ws_cfg, "report_name", "Status Report – RAID PMAR")

    # Fase atual: derivada da aba FASES (não hardcoded)
    detected_phase = _derive_current_phase(wb)
    if detected_phase:
        set_kv(ws_cfg, "current_phase", detected_phase)

    if current_day is not None:
        set_kv(ws_cfg, "current_day", current_day)
    if total_days is not None:
        set_kv(ws_cfg, "total_days", total_days)
    if not keep_existing_progress:
        set_kv(ws_cfg, "progress_percent", 0)
    set_kv(ws_cfg, "pmar_source", _display_path(source_path))
    set_kv(ws_cfg, "last_pmar_import", datetime.now().strftime("%d/%m/%Y %H:%M"))
    ws_cfg.column_dimensions["A"].width = 24
    ws_cfg.column_dimensions["B"].width = 64

    # ── KPIS ────────────────────────────────────────────────────────────────
    ws_kpi = ensure_sheet(wb, "KPIS")
    clear_table_rows(ws_kpi)
    prazo_text = (
        f"Dia {current_day} de {total_days}"
        if current_day is not None and total_days is not None
        else "--"
    )
    prazo_sub = (
        f"{fmt_date(config.get('Start Date'))} a {fmt_date(config.get('End Date'))}"
        if config.get("Start Date") and config.get("End Date")
        else ""
    )
    progress_pct = None
    for row in wb["CONFIG"].iter_rows(min_row=1, max_col=2):
        if row[0].value == "progress_percent":
            progress_pct = row[1].value
            break
    progress_val = f"{progress_pct}%" if progress_pct not in (None, "") else "--"

    n_open = len(open_risks)
    _crit_suffix = f" / {critical_count} crítico{pluralize(critical_count, '', 's')}" if critical_count else ""
    risco_val = f"{n_open} {pluralize(n_open, 'aberto', 'abertos')}{_crit_suffix}"
    risco_nivel = "danger" if critical_count else ("warning" if open_risks else "success")
    saude_val = "Crítico" if critical_count else ("Atenção" if open_risks else "OK")
    saude_sub = f"{high_count} {pluralize(high_count, 'risco alto', 'riscos altos')}" if high_count else ""
    saude_nivel = "danger" if critical_count else ("warning" if open_risks else "success")

    kpis = [
        [1, "Data",        report_dt.strftime("%d/%m/%Y"), "",          "calendar", "success"],
        [2, "Fase Atual",  detected_phase or "--",          "",          "compass",  "success"],
        [3, "Progresso",   progress_val,                   "",          "progress", "success"],
        [4, "Prazo",       prazo_text,                     prazo_sub,   "flag",     "warning"],
        [5, "Risco Atual", risco_val,                      "PMAR RAID", "warning",  risco_nivel],
        [6, "Saúde Geral", saude_val,                      saude_sub,   "heart",    saude_nivel],
    ]
    append_rows(ws_kpi, kpis)
    style_sheet(ws_kpi, {"B": 18, "C": 20, "D": 28})

    # ── RESUMO_EXECUTIVO ─────────────────────────────────────────────────────
    ws_res = ensure_sheet(wb, "RESUMO_EXECUTIVO")
    clear_table_rows(ws_res)
    resumo: list[list] = []
    if open_risks:
        resumo.append([
            1,
            f"{n_open} {pluralize(n_open, 'risco permanece aberto', 'riscos permanecem abertos')}"
            f", sendo {critical_count} crítico{pluralize(critical_count, '', 's')}"
            f" e {high_count} {pluralize(high_count, 'alto', 'altos')}.",
            "andamento",
        ])
        crit = next((r for r in risks if str(clean(r.get("PRIORITY")) or "").lower() == "critical"), None)
        if crit:
            resumo.append([2, f"Risco crítico principal: {crit.get('RISK DESCRIPTION')}", "andamento"])
    else:
        resumo.append([1, "Nenhum risco aberto identificado na PMAR.", "concluido"])
    ordem = len(resumo) + 1
    for r in resolved_risks:
        resumo.append([ordem, f"Risco resolvido: {r.get('RISK DESCRIPTION')}", "concluido"])
        ordem += 1
    append_rows(ws_res, resumo)
    style_sheet(ws_res, {"A": 8, "B": 78, "C": 16})

    # ── PENDENCIAS_CRITICAS ──────────────────────────────────────────────────
    # Coluna data_limite adicionada para expor DUE DATE da PMAR
    ws_pend = ensure_sheet(wb, "PENDENCIAS_CRITICAS")
    clear_sheet(ws_pend)
    pend_headers = [
        "prioridade", "item", "responsaveis", "status", "nivel",
        "id_origem", "categoria", "score", "probabilidade", "impacto",
        "estrategia", "data_limite", "comentarios",
    ]
    ws_pend.append(pend_headers)
    pend_rows: list[list] = []
    for r in open_risks:
        prioridade, nivel = map_priority(r.get("PRIORITY"))
        pend_rows.append([
            prioridade,
            r.get("RISK DESCRIPTION"),
            r.get("OWNER"),
            translate_status(r.get("STATUS")),
            nivel,
            r.get("ID"),
            r.get("CATEGORY"),
            r.get("RISK SCORE"),
            r.get("PROBABILITY"),
            r.get("IMPACT"),
            r.get("STRATEGY"),
            fmt_date(r.get("DUE DATE")),
            r.get("COMMENTS"),
        ])
    append_rows(ws_pend, pend_rows)
    style_sheet(ws_pend, {
        "A": 12, "B": 46, "C": 34, "D": 16, "E": 14,
        "F": 12, "G": 16, "H": 10, "I": 14, "J": 10,
        "K": 16, "L": 14, "M": 50,
    })

    # ── PROXIMAS_ACOES ───────────────────────────────────────────────────────
    ws_acoes = ensure_sheet(wb, "PROXIMAS_ACOES")
    clear_table_rows(ws_acoes)
    acoes: list[list] = []
    for idx, r in enumerate(open_risks, start=1):
        response = clean(r.get("RISK RESPONSE"))
        if response:
            prefix = clean(r.get("ID")) or f"R{idx:03d}"
            acoes.append([idx, f"{prefix}: {response}"])
        else:
            prefix = clean(r.get("ID")) or f"R{idx:03d}"
            acoes.append([idx, f"{prefix}: Definir plano de resposta para '{r.get('RISK DESCRIPTION', '')[:60]}'"])
    append_rows(ws_acoes, acoes)
    style_sheet(ws_acoes, {"A": 8, "B": 82})

    # ── MARCOS: atualiza data de Go-Live e Hypercare ─────────────────────────
    if "MARCOS" in wb.sheetnames and end_dt:
        ws_m = wb["MARCOS"]
        for row in ws_m.iter_rows(min_row=2):
            nome = str(row[1].value or "").lower() if len(row) > 1 else ""
            if "go-live" in nome or "go live" in nome:
                row[2].value = fmt_date(config.get("End Date"))
    if "FASES" in wb.sheetnames and end_dt:
        ws_f = wb["FASES"]
        for row in ws_f.iter_rows(min_row=2):
            nome = str(row[1].value or "").lower() if len(row) > 1 else ""
            if "hypercare" in nome:
                row[3].value = fmt_date(config.get("End Date"))

    # ── RODAPE: apenas campos específicos do milestone ───────────────────────
    # Owner/nome/data do relatório são lidos diretamente do CONFIG pelo frontend
    ws_rod = ensure_sheet(wb, "RODAPE")
    set_kv(ws_rod, "milestone_alvo", "RAID PMAR – Controle de Riscos")
    set_kv(ws_rod, "data_alvo", fmt_date(config.get("End Date")))
    set_kv(ws_rod, "go_live_previsto", fmt_date(config.get("End Date")))
    ws_rod.column_dimensions["A"].width = 24
    ws_rod.column_dimensions["B"].width = 64

    # ── RAID_RISCOS ──────────────────────────────────────────────────────────
    ws_rr = ensure_sheet(wb, "RAID_RISCOS")
    clear_sheet(ws_rr)
    risk_headers = [
        "ID", "CATEGORY", "RISK DESCRIPTION", "RAISED BY", "DATE IDENTIFIED",
        "PROBABILITY", "IMPACT", "RISK SCORE", "PRIORITY", "STRATEGY",
        "RISK RESPONSE", "OWNER", "DUE DATE", "STATUS", "DAYS OPEN",
        "NEXT REVIEW", "COMMENTS",
    ]
    ws_rr.append(risk_headers)
    for r in risks:
        date_cols = {"DATE IDENTIFIED", "DUE DATE", "NEXT REVIEW"}
        ws_rr.append([
            fmt_date(r.get(h)) if h in date_cols else r.get(h)
            for h in risk_headers
        ])
    style_sheet(ws_rr, {
        "A": 10, "B": 15, "C": 48, "D": 16, "E": 16,
        "F": 14, "G": 10, "H": 12, "I": 12, "J": 16,
        "K": 48, "L": 34, "M": 14, "N": 16, "O": 12, "P": 14, "Q": 58,
    })

    # ── RAID opcionais ───────────────────────────────────────────────────────
    optional_specs = [
        (
            "RAID_ACOES", pmar_data["actions"],
            ["ID", "ASSOC. ID", "CATEGORY", "RAISED BY", "CREATE DATE",
             "ACTION DESCRIPTION", "PRIORITY", "ASSIGNED TO", "DUE DATE",
             "STATUS", "% COMPLETE", "DAYS OPEN", "NEXT REVIEW", "COMMENTS"],
        ),
        (
            "RAID_ISSUES", pmar_data["issues"],
            ["ID", "ASSOC. ID", "CATEGORY", "RAISED BY", "CREATE DATE",
             "ISSUE DESCRIPTION", "PRIORITY", "SEVERITY", "RESOLUTION PLAN",
             "ASSIGNED TO", "DUE DATE", "STATUS", "ROOT CAUSE", "DAYS OPEN", "COMMENTS"],
        ),
        (
            "RAID_DECISOES", pmar_data["decisions"],
            ["ID", "ASSOC. ID", "CATEGORY", "DECISION MADE BY", "DATE DECIDED",
             "DECISION TYPE", "DECISION", "RATIONALE", "IMPACT ASSESSMENT",
             "ASSIGNED TO", "DUE DATE", "STATUS", "COMMENTS"],
        ),
    ]
    for sheet_name, rows, headers in optional_specs:
        ws_opt = ensure_sheet(wb, sheet_name)
        clear_sheet(ws_opt)
        ws_opt.append(headers)
        for r in rows:
            ws_opt.append([
                fmt_date(r.get(h)) if "DATE" in h or h in {"DUE DATE", "NEXT REVIEW"} else r.get(h)
                for h in headers
            ])
        style_sheet(ws_opt)

    # ── EQUIPE ───────────────────────────────────────────────────────────────
    ws_eq = ensure_sheet(wb, "EQUIPE")
    clear_sheet(ws_eq)
    ws_eq.append(["nome", "papel", "email"])
    for member in team:
        ws_eq.append([member.get("nome"), member.get("papel"), member.get("email")])
    style_sheet(ws_eq, {"A": 28, "B": 32, "C": 36})

    # ── HISTORICO_IMPORTACAO (acumulativo) ───────────────────────────────────
    ws_hist = ensure_sheet(wb, "HISTORICO_IMPORTACAO")
    _update_historico(ws_hist, {
        "data_importacao":    datetime.now().strftime("%d/%m/%Y %H:%M"),
        "arquivo_origem":     _display_path(source_path),
        "riscos_importados":  len(risks),
        "riscos_abertos":     len(open_risks),
        "riscos_resolvidos":  len(resolved_risks),
        "acoes":              len(pmar_data["actions"]),
        "issues":             len(pmar_data["issues"]),
        "decisoes":           len(pmar_data["decisions"]),
    })
    style_sheet(ws_hist, {"A": 20, "B": 28, "C": 18, "D": 15, "E": 18, "F": 10, "G": 10, "H": 12})

    wb.save(status_path)
    wb.close()

    return {
        "riscos_total":    len(risks),
        "riscos_abertos":  len(open_risks),
        "riscos_resolvidos": len(resolved_risks),
        "acoes":           len(pmar_data["actions"]),
        "issues":          len(pmar_data["issues"]),
        "decisoes":        len(pmar_data["decisions"]),
    }


def _display_path(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(ROOT_DIR.resolve()))
    except Exception:
        return str(path)


def import_pmar(
    pmar_path: Path = DEFAULT_PMAR_PATH,
    status_path: Path = DEFAULT_STATUS_PATH,
) -> dict[str, Any]:
    pmar_data = read_pmar(pmar_path)
    return update_status_workbook(status_path, pmar_data, pmar_path)


def main():
    parser = argparse.ArgumentParser(description="Importa PMAR RAID Log para status_projeto.xlsx")
    parser.add_argument("--pmar",   type=Path, default=DEFAULT_PMAR_PATH, help="Caminho do arquivo PMAR .xlsm")
    parser.add_argument("--status", type=Path, default=DEFAULT_STATUS_PATH, help="Caminho do status_projeto.xlsx")
    args = parser.parse_args()

    result = import_pmar(args.pmar, args.status)
    print("Importação concluída:")
    for key, value in result.items():
        print(f"  {key}: {value}")


if __name__ == "__main__":
    main()
