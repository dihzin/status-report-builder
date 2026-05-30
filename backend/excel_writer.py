import logging
import openpyxl

logger = logging.getLogger(__name__)


def _is_nav_row(row_values):
    v = str(row_values[0] or '').strip()
    return 'Voltar' in v or v.startswith('←')


def _update_key_value(ws, updates):
    """Update values in a key-value sheet (col A = key, col B = value)."""
    for row in ws.iter_rows():
        if not row or row[0].value is None:
            continue
        if len(row) < 2:
            continue
        key = str(row[0].value).strip()
        if key in updates:
            row[1].value = updates[key]


def _rewrite_table(ws, rows_data, columns):
    """Clear data rows (preserve header/nav) and write new rows."""
    first_vals = [c.value for c in ws[1]]
    data_start = 3 if (first_vals and _is_nav_row(first_vals)) else 2

    # Clear existing data rows
    for row in ws.iter_rows(min_row=data_start, max_row=max(ws.max_row, data_start)):
        for cell in row:
            cell.value = None

    # Write new rows
    for r_offset, row_data in enumerate(rows_data):
        row_num = data_start + r_offset
        for c_offset, col_name in enumerate(columns):
            ws.cell(row=row_num, column=c_offset + 1).value = row_data.get(col_name)


def write_excel(filepath, data):
    """Write updated data sections back to the Excel file."""
    try:
        wb = openpyxl.load_workbook(filepath)
    except PermissionError:
        raise Exception("Arquivo Excel está aberto. Feche o arquivo e tente novamente.")
    except Exception as e:
        raise Exception(f"Erro ao abrir Excel: {e}")

    try:
        if 'config' in data and 'CONFIG' in wb.sheetnames:
            _update_key_value(wb['CONFIG'], data['config'])

        if 'rodape' in data and 'RODAPE' in wb.sheetnames:
            _update_key_value(wb['RODAPE'], data['rodape'])

        if 'resumo_executivo' in data and 'RESUMO_EXECUTIVO' in wb.sheetnames:
            _rewrite_table(wb['RESUMO_EXECUTIVO'], data['resumo_executivo'],
                           ['ordem', 'texto', 'status'])

        if 'proximas_acoes' in data and 'PROXIMAS_ACOES' in wb.sheetnames:
            _rewrite_table(wb['PROXIMAS_ACOES'], data['proximas_acoes'],
                           ['ordem', 'texto'])

        if 'pendencias_criticas' in data and 'PENDENCIAS_CRITICAS' in wb.sheetnames:
            _rewrite_table(wb['PENDENCIAS_CRITICAS'], data['pendencias_criticas'],
                           ['prioridade', 'item', 'responsaveis', 'status', 'nivel',
                            'id_origem', 'categoria', 'score', 'probabilidade', 'impacto',
                            'estrategia', 'data_limite', 'comentarios'])

        if 'marcos' in data and 'MARCOS' in wb.sheetnames:
            _rewrite_table(wb['MARCOS'], data['marcos'],
                           ['ordem', 'nome', 'data_alvo', 'status', 'tipo'])

        if 'fases' in data and 'FASES' in wb.sheetnames:
            _rewrite_table(wb['FASES'], data['fases'],
                           ['ordem', 'nome', 'status', 'data_inicio', 'data_alvo', 'destaque'])

        if 'kpis' in data and 'KPIS' in wb.sheetnames:
            _rewrite_table(wb['KPIS'], data['kpis'],
                           ['ordem', 'titulo', 'valor', 'subtitulo', 'tipo', 'nivel'])

        if 'curva_s' in data and 'CURVA_S' in wb.sheetnames:
            _rewrite_table(wb['CURVA_S'], data['curva_s'],
                           ['dia', 'planejado', 'realizado'])

        wb.save(filepath)
        logger.info("Excel salvo com sucesso: %s", filepath)

    except Exception:
        raise
    finally:
        try:
            wb.close()
        except Exception:
            pass
