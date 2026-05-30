"""Formata visualmente o status_projeto.xlsx para UX amigável e idempotente."""
import shutil, sys, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

sys.stdout.reconfigure(encoding='utf-8')

PATH = 'status_projeto.xlsx'
shutil.copy2(PATH, 'status_projeto_backup_fmt.xlsx')

# ── Paleta executiva (neutro + verde institucional) ──────────────────────────
C_HDR_BG       = '143E2A'   # verde escuro institucional
C_HDR_FONT     = 'FFFFFF'
C_ROW_ALT      = 'F4FBF5'   # verde quase branco
C_ROW_WHITE    = 'FFFFFF'
C_TAB_GREEN    = '2A7249'
C_TAB_ORANGE   = 'E65100'
C_TAB_GRAY     = '616161'
C_KV_KEY_BG    = 'F1F8E9'
C_KV_AUTO_BG   = 'FFEBEE'   # campos automáticos (não editar)
C_HINT_BG      = 'FFF9C4'
C_LEGEND_BG    = 'FAFAFA'
C_BORDER       = 'BDBDBD'
C_FONT_DARK    = '212121'
C_FONT_MUTED   = '757575'
C_FONT_AUTO    = 'C62828'


def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def font_style(bold=False, color=C_HDR_FONT, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)


def align_style(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def thin_border():
    s = Side(style='thin', color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def thick_bottom_border():
    thin  = Side(style='thin',   color=C_BORDER)
    thick = Side(style='medium', color='757575')
    return Border(left=thin, right=thin, top=thin, bottom=thick)


def set_header_row(ws, cols_widths):
    """Estiliza linha 1 como cabeçalho e define larguras."""
    for col_idx, (_, width) in enumerate(cols_widths, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill      = fill(C_HDR_BG)
        cell.font      = font_style(bold=True, color=C_HDR_FONT, size=10)
        cell.alignment = align_style('center')
        cell.border    = thick_bottom_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 24


def style_data_rows(ws, start_row, n_cols, alt=C_ROW_ALT, white=C_ROW_WHITE):
    """Alterna cores de fundo nas linhas de dados."""
    for r in range(start_row, ws.max_row + 1):
        bg = alt if r % 2 == 0 else white
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill      = fill(bg)
            cell.font      = Font(size=10, color=C_FONT_DARK)
            cell.alignment = align_style('left', wrap=True)
            cell.border    = thin_border()
        ws.row_dimensions[r].height = 20


def add_comment(cell, text, author="OnePage System"):
    if cell.comment is None:
        cell.comment = Comment(text, author)
    else:
        cell.comment.text += "\n" + text


def ensure_kv_header(ws):
    """Garante que a aba KV tenha o header 'Campo' / 'Valor' na linha 1."""
    if ws.cell(1, 1).value == 'Campo':
        return
    ws.insert_rows(1)
    for col, label in [(1, 'Campo'), (2, 'Valor')]:
        c = ws.cell(1, col)
        c.value     = label
        c.fill      = fill(C_HDR_BG)
        c.font      = font_style(bold=True, size=10)
        c.alignment = align_style('center')
        c.border    = thick_bottom_border()
    ws.row_dimensions[1].height = 24


def style_kv_sheet(ws, tab_color, auto_fields=None, comments=None, legend=None):
    """Formata aba chave-valor com comentários, destaque automático e legenda."""
    auto_fields = auto_fields or set()
    comments = comments or {}
    ws.sheet_properties.tabColor = tab_color
    ensure_kv_header(ws)

    for r in range(2, ws.max_row + 1):
        key  = str(ws.cell(r, 1).value or '')
        auto = key in auto_fields
        bg_k = C_KV_AUTO_BG if auto else C_KV_KEY_BG
        bg_v = 'FFF8E1' if auto else C_ROW_WHITE
        font_k = Font(size=10, bold=True, color=C_FONT_AUTO if auto else '37474F', italic=auto)
        font_v = Font(size=10, color=C_FONT_DARK)

        k = ws.cell(r, 1)
        k.fill      = fill(bg_k)
        k.font      = font_k
        k.alignment = align_style('left')
        k.border    = thin_border()

        v = ws.cell(r, 2)
        v.fill      = fill(bg_v)
        v.font      = font_v
        v.alignment = align_style('left', wrap=True)
        v.border    = thin_border()

        ws.row_dimensions[r].height = 20

        if key in comments:
            add_comment(k, comments[key])
            add_comment(v, comments[key])

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 56
    ws.freeze_panes = 'B2'

    if legend:
        last = ws.max_row + 2
        ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=2)
        leg = ws.cell(last, 1)
        leg.value     = legend
        leg.fill      = fill(C_LEGEND_BG)
        leg.font      = Font(size=9, italic=True, color=C_FONT_MUTED)
        leg.alignment = align_style('left', wrap=True)
        ws.row_dimensions[last].height = 28


def apply_list_validation(ws, col_letter, items, start_row=2):
    """Aplica validação de lista em uma coluna."""
    formula = '"' + ','.join(items) + '"'
    end_row = max(ws.max_row, start_row)
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Escolha um valor da lista."
    dv.errorTitle = "Entrada inválida"
    dv.prompt = "Selecione um valor da lista disponível."
    dv.promptTitle = "Ajuda de preenchimento"
    ws.add_data_validation(dv)
    dv.add(f'{col_letter}{start_row}:{col_letter}{end_row}')


def apply_whole_validation(ws, col_letter, min_val, max_val, start_row=2):
    """Aplica validação de número inteiro em uma coluna."""
    end_row = max(ws.max_row, start_row)
    dv = DataValidation(type="whole", operator="between", formula1=min_val, formula2=max_val, allow_blank=True)
    dv.error = f"Digite um número inteiro entre {min_val} e {max_val}."
    dv.errorTitle = "Valor fora do intervalo"
    dv.prompt = f"Informe um número inteiro de {min_val} a {max_val}."
    dv.promptTitle = "Ajuda de preenchimento"
    ws.add_data_validation(dv)
    dv.add(f'{col_letter}{start_row}:{col_letter}{end_row}')


def add_header_comments(ws, mapping):
    """Adiciona comentários nos headers (linha 1) baseado em mapping {col_index: texto}."""
    for col_idx, text in mapping.items():
        cell = ws.cell(row=1, column=col_idx)
        add_comment(cell, text)


# =============================================================================
wb = openpyxl.load_workbook(PATH)

# ── BRANDING ──────────────────────────────────────────────────────────────────
BRANDING_COMMENTS = {
    'cor_primaria':   'Cor principal da marca (hex). Ex: #2a7249. Usada em acentos e ícones.',
    'cor_secundaria': 'Cor escura da marca (hex). Ex: #1a4f35. Usada em fundos e gradientes.',
    'cor_texto':      'Cor de texto sobre fundo escuro (hex). Ex: #ffffff.',
    'logo_path':      'Caminho relativo do arquivo de logo. Ex: assets/logo.svg.',
}
if 'BRANDING' in wb.sheetnames:
    style_kv_sheet(
        wb['BRANDING'], C_TAB_ORANGE,
        comments=BRANDING_COMMENTS,
        legend='  Dica: As cores devem estar no formato hexadecimal (#RRGGBB). O logo deve ser SVG ou PNG.',
    )

# ── CONFIG ────────────────────────────────────────────────────────────────────
CONFIG_COMMENTS = {
    'logo_path':              'Caminho relativo do logo. Ex: assets/logo.svg',
    'report_title':           'Título do relatório exibido no navegador e capa.',
    'project_name':           'Nome do projeto no topo do slide principal.',
    'project_subtitle':       'Subtítulo curto logo abaixo do nome do projeto.',
    'sponsor':                'Nome do sponsor/cliente (exibido na capa).',
    'pmar_source':            'Caminho do arquivo PMAR/RAID importado.',
    'last_pmar_import':       'Data da última importação PMAR (preenchido automaticamente).',
    'alert_label':            'Texto do alerta no topo (ex: Atenção). Deixe vazio para ocultar.',
    'alert_level':            'warning (amarelo), danger (vermelho), success (verde).',
    'report_date':            'Data do relatório. Formato: DD/MM/AAAA.',
    'current_phase':          'Nome da fase atual (ex: Explorer). Deve constar na aba FASES.',
    'current_day':            'Dia atual do projeto (número inteiro).',
    'total_days':             'Total de dias do cronograma (número inteiro).',
    'progress_percent':         'Percentual geral de avanço (0 a 100).',
    'owner_name':             'Nome do apresentador/owner do relatório.',
    'report_name':            'Nome interno do relatório (exibido no rodapé).',
    'partner_name':           'Nome do parceiro/consultoria (exibido na capa).',
    'presentation_duration':  'Duração da apresentação (ex: 30 minutos).',
    'cover_eyebrow':          'Texto pequeno acima do título da capa (ex: STATUS REPORT DO PROJETO).',
    'cover_main_title':       'Título principal. Use "|" para quebra de linha.',
    'cover_subtitle':         'Subtítulo da capa.',
    'cover_highlight':        'Palavra a ser destacada em itálico no título.',
    'cover_tagline':            'Frase de destaque na parte inferior da capa.',
    'cover_restriction_label': 'Selo de restrição (ex: USO RESTRITO).',
    'cover_footer_left':        'Texto esquerdo do rodapé da capa.',
    'cover_footer_right':       'Texto direito do rodapé da capa.',
}
style_kv_sheet(
    wb['CONFIG'], C_TAB_GREEN,
    auto_fields={'last_pmar_import'},
    comments=CONFIG_COMMENTS,
    legend='  Fundo vermelho = preenchido automaticamente pelo sistema (não editar).   |   Fundo verde = campo editável.',
)

# ── FASES ─────────────────────────────────────────────────────────────────────
ws = wb['FASES']
ws.sheet_properties.tabColor = C_TAB_GREEN
set_header_row(ws, [
    ('ordem', 10), ('nome', 28), ('status', 20), ('data_inicio', 18), ('data_alvo', 18), ('destaque', 12),
])
style_data_rows(ws, 2, 6)
ws.freeze_panes = 'A2'
apply_list_validation(ws, 'C', ['Concluído', 'Em andamento', 'Planejado'])
apply_list_validation(ws, 'F', ['TRUE', 'FALSE'])
add_header_comments(ws, {
    1: 'Número da sequência (1, 2, 3...). Define a ordem de exibição na timeline.',
    2: 'Nome da fase (ex: Prepare, Explorer, Realize).',
    3: "Selecione da lista. 'Concluído' encerra a fase, 'Em andamento' destaca, 'Planejado' deixa tracejado.",
    4: 'Data de início da fase. Formato: DD/MM/AAAA. Opcional — se vazio, o Gantt exibe apenas o marco de término.',
    5: 'Data prevista de término. Formato: DD/MM/AAAA.',
    6: 'TRUE para marcar como fase atual (destaque visual no slide).',
})

# ── KPIS ──────────────────────────────────────────────────────────────────────
ws = wb['KPIS']
ws.sheet_properties.tabColor = C_TAB_GREEN
set_header_row(ws, [
    ('ordem', 8), ('titulo', 24), ('valor', 22), ('subtitulo', 32),
    ('tipo', 16), ('nivel', 14),
])
style_data_rows(ws, 2, 6)
ws.freeze_panes = 'A2'
apply_list_validation(ws, 'E', ['calendar', 'compass', 'progress', 'flag', 'warning', 'heart'])
apply_list_validation(ws, 'F', ['success', 'warning', 'danger', 'gray'])
add_header_comments(ws, {
    1: 'Ordem de exibição na linha de KPIs (da esquerda para a direita).',
    2: 'Nome do indicador (ex: Data, Fase Atual, Progresso).',
    3: 'Texto ou número a ser exibido no card.',
    4: 'Texto complementar opcional (exibido abaixo do valor).',
    5: 'Ícone do card. Escolha da lista: calendar, compass, progress, flag, warning, heart.',
    6: 'Cor do indicador. success=verde, warning=laranja, danger=vermelho, gray=neutro.',
})

# ── RESUMO_EXECUTIVO ──────────────────────────────────────────────────────────
ws = wb['RESUMO_EXECUTIVO']
ws.sheet_properties.tabColor = C_TAB_GREEN
set_header_row(ws, [('ordem', 8), ('texto', 80), ('status', 16)])
style_data_rows(ws, 2, 3)
ws.freeze_panes = 'A2'
apply_list_validation(ws, 'C', ['concluido', 'andamento'])
add_header_comments(ws, {
    1: 'Ordem do item na lista.',
    2: 'Frase do resumo executivo. Use linguagem objetiva e direta.',
    3: 'concluido = item realizado; andamento = item em aberto/risco.',
})

# ── PENDENCIAS_CRITICAS ─────────────────────────────────────────────────────
ws = wb['PENDENCIAS_CRITICAS']
ws.sheet_properties.tabColor = C_TAB_GREEN
hdr = [c.value for c in ws[1]]
if 'data_limite' not in hdr:
    ws.insert_cols(12)
    ws.cell(1, 12).value = 'data_limite'
set_header_row(ws, [
    ('prioridade', 12), ('item', 50), ('responsaveis', 30), ('status', 20),
    ('nivel', 14), ('id_origem', 14), ('categoria', 18), ('score', 10),
    ('probabilidade', 14), ('impacto', 10), ('estrategia', 18),
    ('data_limite', 18), ('comentarios', 50),
])
style_data_rows(ws, 2, 13)
ws.freeze_panes = 'B2'
apply_list_validation(ws, 'A', ['P1', 'P2', 'P3', 'P4'])
apply_list_validation(ws, 'D', ['Atrasado', 'Em atenção', 'No prazo'])
apply_list_validation(ws, 'E', ['danger', 'warning', 'success', 'gray'])
add_header_comments(ws, {
    1: 'P1 (crítico), P2 (alto), P3 (médio), P4 (baixo).',
    2: 'Descrição resumida do risco ou pendência. Seja claro e objetivo.',
    3: 'Nomes separados por barra (/). Ex: Rodrigo / Bruna.',
    4: 'Status atual: Atrasado, Em atenção, No prazo.',
    5: 'Nível de severidade. danger=vermelho, warning=laranja, success=verde, gray=neutro.',
    6: 'ID no sistema de origem (ex: R001).',
    7: 'Classificação (ex: Schedule, External, Finance, Scope).',
    8: 'Score de risco (número inteiro).',
    9: 'Probabilidade de 0 a 1 (ex: 0.7).',
    10: 'Impacto de 1 a 5 (número inteiro).',
    11: 'Estratégia (ex: Mitigate, Accept, Transfer, Avoid).',
    12: 'Prazo para resolução. Formato: DD/MM/AAAA.',
    13: 'Detalhes adicionais, contexto ou próximos passos.',
})

# ── PROXIMAS_ACOES ────────────────────────────────────────────────────────────
ws = wb['PROXIMAS_ACOES']
ws.sheet_properties.tabColor = C_TAB_GREEN
set_header_row(ws, [('ordem', 8), ('texto', 80)])
style_data_rows(ws, 2, 2)
ws.freeze_panes = 'A2'
add_header_comments(ws, {
    1: 'Ordem da ação na lista.',
    2: 'Descrição da próxima ação. Seja específico: quem faz o quê e até quando.',
})

# ── CURVA_S ───────────────────────────────────────────────────────────────────
ws = wb['CURVA_S']
ws.sheet_properties.tabColor = C_TAB_GREEN
set_header_row(ws, [
    ('dia', 12), ('planejado', 18), ('realizado', 18),
])
style_data_rows(ws, 2, 3)
ws.freeze_panes = 'A2'
apply_whole_validation(ws, 'B', 0, 100)
apply_whole_validation(ws, 'C', 0, 100)
add_header_comments(ws, {
    1: 'Dia do projeto (número inteiro). Ex: 1, 7, 14...',
    2: '% planejado de conclusão acumulada (0 a 100).',
    3: '% realizado de conclusão acumulada (0 a 100).',
})

# ── MARCOS ────────────────────────────────────────────────────────────────────
ws = wb['MARCOS']
ws.sheet_properties.tabColor = C_TAB_GREEN
set_header_row(ws, [
    ('ordem', 10), ('nome', 40), ('data_alvo', 18), ('status', 20), ('tipo', 16),
])
style_data_rows(ws, 2, 5)
ws.freeze_panes = 'A2'
apply_list_validation(ws, 'D', ['Concluído', 'Em andamento', 'Planejado'])
apply_list_validation(ws, 'E', ['check', 'rocket', 'gear', 'star'])
add_header_comments(ws, {
    1: 'Ordem do marco na timeline.',
    2: 'Nome do marco (ex: Kick-off concluído, Gate Explorer → Realize).',
    3: 'Data prevista de conclusão. Formato: DD/MM/AAAA.',
    4: 'Status atual do marco. Escolha da lista.',
    5: 'Ícone visual: check, rocket, gear, star.',
})

# ── RODAPE ────────────────────────────────────────────────────────────────────
RODAPE_COMMENTS = {
    'milestone_alvo': 'Próximo marco a ser alcançado (ex: Explorer → Realize).',
    'data_alvo':      'Data alvo do próximo marco. Formato: DD/MM/AAAA.',
    'go_live_previsto': 'Data prevista do Go-Live final. Formato: DD/MM/AAAA.',
}
style_kv_sheet(
    wb['RODAPE'], C_TAB_GREEN,
    comments=RODAPE_COMMENTS,
    legend='  Campos do rodapé do slide principal. Vincule-se aos marcos e datas da aba MARCOS.',
)

# ── GANTT_TAREFAS ────────────────────────────────────────────────────────────
if 'GANTT_TAREFAS' in wb.sheetnames:
    ws = wb['GANTT_TAREFAS']
    ws.sheet_properties.tabColor = C_TAB_ORANGE
    set_header_row(ws, [
        ('id', 10), ('parent_id', 12), ('nome', 36), ('inicio', 14), ('fim', 14),
        ('progresso', 12), ('status', 14), ('owner', 22), ('dependencias', 18),
    ])
    style_data_rows(ws, 2, 9)
    ws.freeze_panes = 'A2'
    apply_list_validation(ws, 'G', ['Very High', 'High', 'Medium', 'Low', 'Concluído', 'Em andamento', 'Planejado', 'Atrasado'])
    add_header_comments(ws, {
        1: 'Identificador único da tarefa. Não repetir.',
        2: 'ID da tarefa pai (para hierarquia). Deixe vazio para tarefa raiz.',
        3: 'Nome da tarefa ou fase.',
        4: 'Data de início. Formato: DD/MM/AAAA.',
        5: 'Data de término. Formato: DD/MM/AAAA.',
        6: 'Percentual concluído (0 a 100).',
        7: 'Prioridade / status. Escolha da lista.',
        8: 'Nome do responsável pela tarefa.',
        9: 'IDs das tarefas predecessoras, separados por vírgula.',
    })

# ── GANTT_MARCOS ─────────────────────────────────────────────────────────────
if 'GANTT_MARCOS' in wb.sheetnames:
    ws = wb['GANTT_MARCOS']
    ws.sheet_properties.tabColor = C_TAB_ORANGE
    set_header_row(ws, [
        ('id', 10), ('nome', 36), ('data', 14), ('status', 16), ('tipo', 12),
    ])
    style_data_rows(ws, 2, 5)
    ws.freeze_panes = 'A2'
    apply_list_validation(ws, 'D', ['Concluído', 'Em andamento', 'Planejado', 'Atrasado'])
    apply_list_validation(ws, 'E', ['check', 'rocket', 'gear', 'star'])
    add_header_comments(ws, {
        1: 'Identificador único do marco.',
        2: 'Nome do marco (ex: Go-Live, Gate de decisão).',
        3: 'Data do marco. Formato: DD/MM/AAAA.',
        4: 'Status atual. Escolha da lista.',
        5: 'Ícone visual: check, rocket, gear, star.',
    })

# ── GANTT_CONFIG ──────────────────────────────────────────────────────────────
if 'GANTT_CONFIG' in wb.sheetnames:
    style_kv_sheet(
        wb['GANTT_CONFIG'], C_TAB_ORANGE,
        legend='  Configurações de visualização do Gantt no slide 3.',
    )

# ── Reordena abas ─────────────────────────────────────────────────────────────
ORDER = [
    'BRANDING', 'CONFIG', 'PRESENTATION_CONFIG', 'FASES', 'KPIS', 'RESUMO_EXECUTIVO',
    'PENDENCIAS_CRITICAS', 'PROXIMAS_ACOES', 'CURVA_S', 'MARCOS', 'RODAPE',
    'GANTT_TAREFAS', 'GANTT_MARCOS', 'GANTT_CONFIG',
]
for i, name in enumerate(ORDER):
    if name in wb.sheetnames:
        current = wb.sheetnames.index(name)
        if current != i:
            wb.move_sheet(name, offset=i - current)

wb.save(PATH)
wb.close()
print('Formatacao concluida com sucesso.')
