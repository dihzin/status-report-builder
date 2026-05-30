"""Cria (ou recria) a aba CAPA no status_projeto.xlsx com menu de navegação."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "status_projeto.xlsx"

wb = openpyxl.load_workbook(str(XLSX))

# ── Remove CAPA anterior ───────────────────────────────────────────────────
if "CAPA" in wb.sheetnames:
    del wb["CAPA"]

ws = wb.create_sheet("CAPA", 0)
ws.sheet_view.showGridLines = False

# ── Larguras de colunas ───────────────────────────────────────────────────
# A=margem | B=col1 | C=gap | D=col2 | E=gap | F=col3 | G=margem
for col, w in {"A": 3, "B": 35, "C": 2, "D": 35, "E": 2, "F": 35, "G": 3}.items():
    ws.column_dimensions[col].width = w

# ── Paleta ────────────────────────────────────────────────────────────────
BLUE   = "003873"
ORANGE = "F15921"
WHITE  = "FFFFFF"
BG     = "F4F6FB"
CARD   = "FFFFFF"
GRAY   = "6B7280"
ACCENT = "A0B4D6"

# Cores de categoria — cada bloco tem a sua
CAT_COLORS = {
    "header":   ("EEF2FA", "003873"),   # Cabeçalho  → azul
    "timeline": ("FFF3EE", "C04A10"),   # Timeline   → laranja
    "kpis":     ("EDFAF3", "1A6638"),   # KPIs       → verde
    "middle":   ("FFF8F0", "A05000"),   # Painel ctr → âmbar
    "curvas":   ("EEF6FF", "1B4F8A"),   # Curva S    → azul médio
    "marcos":   ("F5F0FF", "5B2D9E"),   # Marcos     → violeta
    "rodape":   ("F0F5F0", "2C5F3A"),   # Rodapé     → verde escuro
    "gantt":    ("FFF0F5", "8B1A3A"),   # Gantt      → vinho
    "config":   ("F4F4F4", "444444"),   # Config     → cinza
}

# ── Helpers ───────────────────────────────────────────────────────────────
def fill(c):
    return PatternFill("solid", fgColor=c)

def fnt(color=WHITE, size=11, bold=False, italic=False, underline=None):
    return Font(name="Calibri", size=size, bold=bold, italic=italic,
                color=color, underline=underline)

def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def card_border(color="D1D5DB"):
    s = Side(style="thin", color=color)
    return Border(top=s, bottom=s, left=s, right=s)

def set_bg(row_s, row_e, col_s, col_e, color):
    for r in range(row_s, row_e + 1):
        for c in range(col_s, col_e + 1):
            ws.cell(r, c).fill = fill(color)

# ── Fundo geral ───────────────────────────────────────────────────────────
set_bg(1, 80, 1, 7, BG)
for i in range(1, 80):
    ws.row_dimensions[i].height = 16

# ══════════════════════════════════════════════════════════════════════════
# HEADER  (linhas 1-8)
# ══════════════════════════════════════════════════════════════════════════
for r in (1, 8):
    ws.row_dimensions[r].height = 8
for r in range(2, 8):
    ws.row_dimensions[r].height = 22
    for c in range(2, 7):
        ws.cell(r, c).fill = fill(BLUE)

# Acento laranja (linha 7)
ws.row_dimensions[7].height = 5
for c in range(2, 7):
    ws.cell(7, c).fill = fill(ORANGE)

# Título
ws.merge_cells("B3:F4")
cell = ws["B3"]
cell.value = "   Status Report · EDF Power"
cell.font = fnt(WHITE, 20, bold=True)
cell.alignment = aln("left", "center")
cell.fill = fill(BLUE)

# Subtítulo
ws.merge_cells("B5:F5")
cell = ws["B5"]
cell.value = "   EDFP1 - 106024.1.1  |  Clique em uma guia para editá-la diretamente"
cell.font = fnt(ACCENT, 10, italic=True)
cell.alignment = aln("left", "center")
cell.fill = fill(BLUE)

# ══════════════════════════════════════════════════════════════════════════
# Funções de bloco
# ══════════════════════════════════════════════════════════════════════════
_current_row = [9]   # ponteiro mutável para linha atual

def spacer(h=8):
    r = _current_row[0]
    ws.row_dimensions[r].height = h
    set_bg(r, r, 1, 7, BG)
    _current_row[0] += 1
    return r

def section_header(label, cat_key, number):
    """Faixa de categoria com número de bloco e cor própria."""
    bg, fg = CAT_COLORS[cat_key]
    r = _current_row[0]
    ws.row_dimensions[r].height = 24
    ws.merge_cells(f"B{r}:F{r}")
    cell = ws.cell(r, 2)
    cell.value = f"  {number}   {label}"
    cell.font = fnt(fg, 9, bold=True)
    cell.fill = fill(bg)
    cell.alignment = aln("left", "center")
    # borda inferior colorida
    s_bot = Side(style="medium", color=fg)
    s_thin = Side(style="thin", color="D1D5DB")
    for c in range(2, 7):
        cl = ws.cell(r, c)
        cl.fill = fill(bg)
        cl.border = Border(bottom=s_bot, top=s_thin, left=s_thin, right=s_thin)
    _current_row[0] += 1
    return r

def btn_row(cols_data):
    """
    Adiciona uma linha de 1, 2 ou 3 botões.
    cols_data = list of (col_letter, sheet_name, label, desc)  onde col_letter in B/D/F
    """
    COL = {"B": 2, "D": 4, "F": 6}
    r = _current_row[0]
    ws.row_dimensions[r].height = 28

    used_cols = {COL[cd[0]] for cd in cols_data}
    # fundo BG nas colunas vazias da linha
    for c in range(2, 7):
        if c not in used_cols:
            ws.cell(r, c).fill = fill(BG)

    has_desc = any(len(cd) > 3 and cd[3] for cd in cols_data)
    r2 = r + 1
    if has_desc:
        ws.row_dimensions[r2].height = 15

    for cd in cols_data:
        col_letter, sheet_name, label = cd[0], cd[1], cd[2]
        desc = cd[3] if len(cd) > 3 else ""
        col = COL[col_letter]

        cell = ws.cell(r, col)
        cell.value = f"   {label}"
        cell.hyperlink = f"#{sheet_name}!A1"
        cell.font = Font(name="Calibri", size=10, bold=True, color=BLUE)
        cell.fill = fill(CARD)
        cell.alignment = aln("left", "center")
        cell.border = card_border()

        if desc:
            d = ws.cell(r2, col)
            d.value = f"   {desc}"
            d.font = fnt(GRAY, 8, italic=True)
            d.fill = fill(CARD)
            d.alignment = aln("left", "center")
            s = Side(style="thin", color="D1D5DB")
            d.border = Border(left=s, right=s, bottom=s)

    _current_row[0] = r2 + 1 if has_desc else r + 1
    return r

# ══════════════════════════════════════════════════════════════════════════
# BLOCO 1 — Cabeçalho do slide
# Campos: título, subtítulo, fase, data, alerta, Plano %, dia atual
# ══════════════════════════════════════════════════════════════════════════
spacer(10)
section_header("CABEÇALHO DO SLIDE", "header", "1")
btn_row([
    ("B", "CONFIG", "📋  CONFIG",
     "Título, fase, data, alerta, Plano %, dia atual, owner…"),
])
spacer(6)

# ══════════════════════════════════════════════════════════════════════════
# BLOCO 2 — Timeline de Fases
# Faixa horizontal no topo do slide com as fases do projeto
# ══════════════════════════════════════════════════════════════════════════
section_header("TIMELINE DE FASES", "timeline", "2")
btn_row([
    ("B", "FASES", "📍  FASES",
     "Fases do projeto: nome, status (Concluído / Em andamento / Planejado), datas"),
])
spacer(6)

# ══════════════════════════════════════════════════════════════════════════
# BLOCO 3 — KPIs
# Cards de indicadores abaixo da timeline
# ══════════════════════════════════════════════════════════════════════════
section_header("KPIs — INDICADORES-CHAVE", "kpis", "3")
btn_row([
    ("B", "KPIS", "📊  KPIS",
     "Cards de KPI: título, valor, subtítulo, ícone e nível (success/warning/danger)"),
])
spacer(6)

# ══════════════════════════════════════════════════════════════════════════
# BLOCO 4 — Painel central (3 colunas lado a lado)
# ══════════════════════════════════════════════════════════════════════════
section_header("PAINEL CENTRAL  (3 colunas)", "middle", "4")
btn_row([
    ("B", "RESUMO_EXECUTIVO",   "📄  RESUMO EXECUTIVO",
     "Bullets do resumo executivo e respectivo status"),
    ("D", "PENDENCIAS_CRITICAS","⚠️  PENDÊNCIAS CRÍTICAS",
     "Riscos e pendências: prioridade, item, responsável, status"),
    ("F", "PROXIMAS_ACOES",     "✅  PRÓXIMAS AÇÕES",
     "Lista de ações planejadas para o próximo período"),
])
spacer(6)

# ══════════════════════════════════════════════════════════════════════════
# BLOCO 5 — Curva S
# Gráfico de avanço planejado vs realizado
# ══════════════════════════════════════════════════════════════════════════
section_header("CURVA S — AVANÇO PLANEJADO × REALIZADO", "curvas", "5")
btn_row([
    ("B", "CURVA_S", "📈  CURVA S",
     "Dados do gráfico: dia, % planejado, % realizado"),
])
spacer(6)

# ══════════════════════════════════════════════════════════════════════════
# BLOCO 6 — Marcos e Datas-Alvo
# Painel lateral direito no slide principal
# ══════════════════════════════════════════════════════════════════════════
section_header("MARCOS E DATAS-ALVO", "marcos", "6")
btn_row([
    ("B", "MARCOS", "🏁  MARCOS",
     "Marcos do projeto: nome, data-alvo, status e tipo de ícone"),
])
spacer(6)

# ══════════════════════════════════════════════════════════════════════════
# BLOCO 7 — Rodapé do slide
# Faixa inferior com milestone alvo, data e Go-Live
# ══════════════════════════════════════════════════════════════════════════
section_header("RODAPÉ DO SLIDE", "rodape", "7")
btn_row([
    ("B", "RODAPE", "📌  RODAPÉ",
     "Milestone alvo, data-alvo e previsão de Go-Live"),
])
spacer(6)

# ══════════════════════════════════════════════════════════════════════════
# BLOCO 8 — Cronograma Gantt  (Slide 3)
# ══════════════════════════════════════════════════════════════════════════
section_header("CRONOGRAMA GANTT  (Slide 3)", "gantt", "8")
btn_row([
    ("B", "GANTT_TAREFAS", "🗂️  GANTT · TAREFAS",
     "Tarefas do cronograma: id, pai, nome, início, fim, progresso, owner"),
    ("D", "GANTT_MARCOS",  "📌  GANTT · MARCOS",
     "Marcos do Gantt: nome, data, status e tipo"),
    ("F", "GANTT_CONFIG",  "⚙️  GANTT · CONFIG",
     "Escala de tempo, janela de exibição, linhas de progresso e hoje"),
])
spacer(6)

# ══════════════════════════════════════════════════════════════════════════
# BLOCO 9 — Identidade Visual & Configurações Avançadas
# ══════════════════════════════════════════════════════════════════════════
section_header("IDENTIDADE VISUAL & CONFIGURAÇÕES AVANÇADAS", "config", "9")
btn_row([
    ("B", "BRANDING",            "🎨  BRANDING",
     "Cores primária/secundária e caminho do logo"),
    ("D", "PRESENTATION_CONFIG", "🖥️  APRESENTAÇÃO CONFIG",
     "Fontes, tamanhos, cores de alerta e paleta do gráfico"),
])
spacer(6)

# ══════════════════════════════════════════════════════════════════════════
# NOTA DE RODAPÉ DA CAPA
# ══════════════════════════════════════════════════════════════════════════
r_note = _current_row[0]
ws.row_dimensions[r_note].height = 20
ws.merge_cells(f"B{r_note}:F{r_note}")
note = ws.cell(r_note, 2)
note.value = (
    "ℹ️   Fundo verde = preencher manualmente   |   "
    "Fundo vermelho = preenchido automaticamente pelo sistema (não editar)"
)
note.font = fnt(GRAY, 8.5, italic=True)
note.fill = fill(BG)
note.alignment = aln("center", "center")

# ══════════════════════════════════════════════════════════════════════════
# LINK "← Capa" em coluna lateral de cada aba (sem inserir linhas)
# ══════════════════════════════════════════════════════════════════════════
back_font = Font(name="Calibri", size=9, bold=True, color=ORANGE, underline="single")
back_fill = PatternFill("solid", fgColor="FFF8F5")
back_aln  = Alignment(horizontal="left", vertical="center")
side      = Side(style="thin", color="F15921")
back_brd  = Border(left=side, right=side, top=side, bottom=side)

for sheet_name in wb.sheetnames:
    if sheet_name == "CAPA":
        continue
    sh = wb[sheet_name]
    max_col = sh.max_column or 1
    link_col = max_col + 2

    # Limpa link anterior nesta mesma coluna (se existir)
    prev = sh.cell(1, link_col)
    if prev.value and "Capa" in str(prev.value):
        prev.value = None
        prev.hyperlink = None

    c = sh.cell(1, link_col)
    c.value = "← Capa"
    c.hyperlink = "#CAPA!A1"
    c.font = back_font
    c.fill = back_fill
    c.alignment = back_aln
    c.border = back_brd

wb.save(str(XLSX))
print("OK  CAPA criada com sucesso em", XLSX)
