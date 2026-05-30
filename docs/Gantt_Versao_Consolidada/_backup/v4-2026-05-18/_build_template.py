"""Gera os templates Excel para o Gantt Chart Interativo (PT-BR / ES / EN)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date, timedelta
import os

OUT_DIR = os.path.dirname(__file__)

# ============================================================
# Estilos compartilhados
# ============================================================
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="Calibri", size=11)
CELL_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)
BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

# ============================================================
# Traducoes
# ============================================================
TR = {
    "pt": {
        "lang_full": "pt-BR",
        "file_suffix": "pt",
        "sheet_data": "Dados",
        "sheet_instructions": "Instruções",
        "sheet_status": "Glossário Status",
        "sheet_completion": "% Completion",
        "ws2_title": "Como preencher este template — Gantt Chart Interativo",
        "ws2_subtitle": "Preencha a aba 'Dados' e importe este arquivo no gantt.html via botão Importar.",
        "header_required": "Obrigatório",
        "header_type": "Tipo",
        "header_column": "Coluna",
        "header_desc": "Descrição e regras",
        "yes": "Sim",
        "no": "Não",
        "rules_title": "Regras gerais e dicas",
        "status_ref_title": "Status (prioridade) — referência visual",
        "status_when": "Quando usar",
        "status_color": "Cor",
        "status_value": "Valor",
        "comp_title": "Campo 'completion' — lógica detalhada",
        "comp_intro": ("Percentual de conclusão da tarefa, no intervalo 0 a 100. Esta aba explica como o valor é "
                       "interpretado pelo Gantt, como fases agregam, e que critério usar ao preencher."),
        "comp_s1": "1. Como o app renderiza visualmente",
        "comp_s1_text": (
            "A barra da tarefa é desenhada em duas camadas sobrepostas:\n"
            "  • Uma barra de fundo (cor clara, 30% de opacidade) que representa a duração total (start → end).\n"
            "  • Uma barra de progresso (cor cheia) que ocupa exatamente 'completion'% da largura da barra de fundo.\n"
            "Ou seja: completion=0 → só fundo claro;  completion=50 → metade preenchida;  completion=100 → barra totalmente cheia."),
        "comp_visual_title": "Exemplo visual (largura da barra = % preenchido):",
        "comp_s2": "2. Marcos (milestone=TRUE) ignoram este campo",
        "comp_s2_text": (
            "Marcos são pontos no tempo, não períodos — por isso a noção de 'conclusão parcial' não se aplica.\n"
            "O app renderiza marco como um losango cheio, independentemente do valor em 'completion'.\n"
            "Convenção: deixe completion=0 nos marcos."),
        "comp_who_title": "3. Quem preenche e quem o app calcula",
        "comp_who_intro": "Há 3 tipos de linhas. A regra de ouro: você preenche as FOLHAS, o app calcula os RESUMOS.",
        "comp_who_headers": ["Tipo de linha", "Quem decide o %?", "Exemplo no projeto padrão"],
        "comp_who_rows": [
            ["Folha (sem filhos)", "Você preenche manualmente (0-100)", "Wireframes = 80"],
            ["Resumo (tem filhos)", "App calcula. NÃO preencha — será sobrescrito.", "Design = média das folhas"],
            ["Marco (milestone=TRUE)", "Deixe 0. O app ignora no cálculo agregado.", "Go-Live"],
        ],
        "comp_s3": "4. Fórmula recursiva (vale para qualquer profundidade)",
        "comp_s3_text": (
            "Para cada nó-resumo (em qualquer nível: projeto, workstream, fase, sub-fase), o app:\n"
            "  1. Caminha em TODA a árvore descendente coletando as FOLHAS\n"
            "  2. Descarta as que são marcos (milestone=TRUE)\n"
            "  3. Calcula a MÉDIA SIMPLES dos % dessas folhas\n\n"
            "Pseudocódigo:\n"
            "    função calcular_pct(nó):\n"
            "        se nó NÃO tem filhos:           ← é folha\n"
            "            retorna nó.completion\n"
            "        senão:\n"
            "            folhas = todas as descendentes-folha (recursão)\n"
            "            válidas = folhas onde milestone == FALSE\n"
            "            se válidas vazias:\n"
            "                retorna 0\n"
            "            retorna SOMA(válidas[i].completion) / CONTAGEM(válidas)\n\n"
            "Em notação matemática:\n"
            "    %_resumo = Σ(%_folhas_não_marco) / N_folhas_não_marco"),
        "comp_ex_title": "5. Exemplo trabalhado completo — 3 níveis, 13 folhas",
        "comp_ex_intro": ("Considere o dataset padrão (Projeto Exemplo). Veja o cálculo bottom-up — "
                          "cada linha mostra o nível, tipo, fórmula aplicada e resultado:"),
        "comp_ex_headers": ["Nó", "Nível", "Tipo", "Cálculo", "Resultado"],
        "comp_ex_rows": [
            ["Requirements gathering",    "2", "folha",  "—", "50"],
            ["Requirements gathering 2",  "2", "folha",  "—", "55"],
            ["Stakeholder workshop",      "2", "folha",  "—", "75"],
            ["Story boarding",            "2", "folha",  "—", "80"],
            ["Initiation complete",       "2", "marco",  "(ignorado)", "—"],
            ["▸ Initiation",              "1", "resumo", "(50+55+75+80)/4", "65%"],
            ["E2E data solution design",  "2", "folha",  "—", "100"],
            ["Wireframes",                "2", "folha",  "—", "80"],
            ["Prototyping",               "2", "folha",  "—", "40"],
            ["Design complete",           "2", "marco",  "(ignorado)", "—"],
            ["▸ Design",                  "1", "resumo", "(100+80+40)/3", "73%"],
            ["ETL",                       "2", "folha",  "—", "15"],
            ["Data modelling",            "2", "folha",  "—", "40"],
            ["Measures & KPIs",           "2", "folha",  "—", "50"],
            ["Visuals",                   "2", "folha",  "—", "25"],
            ["Implementation complete",   "2", "marco",  "(ignorado)", "—"],
            ["▸ Implementation",          "1", "resumo", "(15+40+50+25)/4", "33%"],
            ["UAT",                       "2", "folha",  "—", "0"],
            ["Bug fixes",                 "2", "folha",  "—", "0"],
            ["Go-Live",                   "2", "marco",  "(ignorado)", "—"],
            ["▸ Testing",                 "1", "resumo", "(0+0)/2", "0%"],
            ["■ PROJETO EXEMPLO",         "0", "raiz",   "soma das 13 folhas / 13 = 610/13", "47%"],
        ],
        "comp_ex_note": ("Verificação do total da raiz: 50+55+75+80 + 100+80+40 + 15+40+50+25 + 0+0 = 610. "
                          "610 ÷ 13 = 46,92% → arredondado para 47%."),
        "comp_warning_title": "6. ATENÇÃO: média das FOLHAS, não média das fases",
        "comp_warning_body": ("O % da raiz NÃO é a média das fases — é a média das folhas DIRETAMENTE. "
                              "Fases com número diferente de tarefas pesam diferente nos dois métodos:"),
        "comp_warning_headers": ["Método de cálculo", "Fórmula", "Resultado"],
        "comp_warning_rows": [
            ["Média das 4 fases (ERRADO)",   "(65 + 73 + 33 + 0) / 4",   "43%"],
            ["Média das 13 folhas (CORRETO)", "610 / 13",                "47%"],
        ],
        "comp_warning_explain": ("A diferença existe porque Testing (só 2 folhas, ambas 0%) puxa MENOS o resultado "
                                  "quando olhamos folha por folha. O app usa média de folhas — escolha consciente. "
                                  "Se quiser média ponderada por duração ou esforço, calcule externamente "
                                  "(o modelo não tem coluna de peso)."),
        "comp_s4": "4. Critério sugerido (rubrica)",
        "comp_s4_text": ("Use a rubrica abaixo para consistência. Trabalhe em múltiplos de 25 — "
                          "precisão maior raramente reflete realidade observável."),
        "comp_rubric_headers": ["%", "Significado", "Quando usar"],
        "comp_rubric_rows": [
            ["0",   "Não iniciada",   "Tarefa ainda não começou."],
            ["25",  "Iniciada",       "Trabalho começou. Entregável parcial."],
            ["50",  "A meio caminho", "Metade do esforço gasto. Entregável existe."],
            ["75",  "Quase pronta",   "Revisada ao menos uma vez. Ajustes finais."],
            ["100", "Concluída",      "Entregável aceito pelo responsável."],
        ],
        "comp_s5": "5. Pegadinhas comuns",
        "comp_gotchas": [
            ("Não use porcentagem temporal (dias decorridos / duração total).",
             "Esse cálculo não diz nada sobre progresso real do entregável."),
            ("Não preencha 'completion' na linha da fase.",
             "É derivada automaticamente pelo app."),
            ("Não use valores fracionários (37,5%).",
             "O app aceita apenas inteiros."),
            ("100% ≠ 'aceito pelo cliente'.",
             "Aceite formal normalmente é um marco separado."),
            ("Cuidado ao virar tarefa em marco com completion > 0.",
             "O valor será ignorado; zere manualmente por higiene."),
            ("100% com 'end' no futuro é suspeito.",
             "Atualize 'end' para a data real de término."),
        ],
        "comp_summary_label": "6. Resumo de 1 linha",
        "comp_summary": ("'completion' é PERCENTUAL DE ENTREGÁVEL produzido (não tempo decorrido), preenchido "
                          "pela equipe em múltiplos de 25, ignorado em marcos, agregado por média simples na linha da fase."),
        "columns": [
            ("id", 8, "Identificador único da tarefa. Inteiro sequencial. Obrigatório para referência de dependências."),
            ("parent_id", 12, "ID da tarefa pai. Vazio = raiz. Define hierarquia N níveis (projeto > fase > task > sub-task...)."),
            ("phase", 22, "[Legado] Nome da fase. Só usado se parent_id estiver vazio. Recomendado: deixe vazio e use parent_id."),
            ("task", 38, "Descrição da tarefa que aparece na barra do Gantt."),
            ("milestone", 12, "TRUE = marco (losango), FALSE = barra normal."),
            ("start", 14, "Data de início. Formato DD/MM/AAAA. Em tarefas-resumo, deixe vazio (calculado dos filhos)."),
            ("end", 14, "Data de término. Formato DD/MM/AAAA. Em tarefas-resumo, deixe vazio."),
            ("completion", 12, "Percentual concluído (0 a 100). Em resumo, calculado dos filhos. Ver aba '% Completion'."),
            ("dependencies", 16, "IDs predecessores separados por vírgula. Ex: 2,3"),
            ("assignee", 18, "Responsável pela tarefa (texto livre)."),
            ("status", 14, "Prioridade: Very High | High | Medium | Low | (vazio)."),
        ],
        "details": [
            ("id", "Inteiro", "Sim", "Identificador único. Use números sequenciais. Referência usada por 'dependencies' e 'parent_id'."),
            ("parent_id", "Inteiro", "Não", "ID da tarefa pai. Vazio = nó raiz. Define hierarquia: Projeto > Workstream > Fase > Sub-fase > Task. Profundidade ilimitada. Ciclos são detectados e impedidos."),
            ("phase", "Texto", "Não", "[Legado] Se parent_id estiver vazio mas phase preenchida, o app cria uma fase-resumo automaticamente. Recomendado: deixe vazio e use parent_id explicitamente."),
            ("task", "Texto", "Sim", "Descrição que aparece na barra. Idealmente até 60 caracteres."),
            ("milestone", "TRUE/FALSE", "Não", "TRUE = losango. FALSE = barra normal. Para marcos, start = end."),
            ("start", "Data DD/MM/AAAA", "Sim", "Data de início. Aceita formato data nativo do Excel."),
            ("end", "Data DD/MM/AAAA", "Sim", "Data de término. Deve ser ≥ start."),
            ("completion", "Inteiro 0-100", "Não", "Percentual de entregável produzido. Ver aba '% Completion' para regras detalhadas."),
            ("dependencies", "Texto", "Não", "IDs das predecessoras separados por vírgula. Ex.: '2,3'."),
            ("assignee", "Texto", "Não", "Responsável (texto livre). Aparece no tooltip."),
            ("status", "Lista", "Não", "Very High / High / Medium / Low / (vazio). Define a cor da barra."),
        ],
        "rules": [
            "1. Não deixe linhas em branco entre tarefas — o importador para na primeira vazia.",
            "2. IDs não precisam ser sequenciais sem buracos, mas devem ser únicos e estáveis.",
            "3. Datas: aceita tanto data Excel quanto texto DD/MM/AAAA.",
            "4. Para criar uma fase nova, basta digitar um novo nome em 'phase'.",
            "5. Marcos representam pontos no tempo — sempre start = end.",
            "6. Dependências circulares geram setas inconsistentes. Evite.",
            "7. Status em branco = cinza neutro (útil para marcos genéricos).",
            "8. Salve, abra gantt.html, clique em Importar, escolha 'Substituir todos'.",
            "9. Você pode misturar este Excel com edição inline depois.",
            "10. Para colaboração via Git, exporte como CSV.",
        ],
        "status_ref": [
            ("Very High", "B91C1C", "Crítico. Caminho crítico ou bloqueador absoluto."),
            ("High",      "EA580C", "Alta prioridade. Atrasos têm impacto significativo."),
            ("Medium",    "CA8A04", "Padrão para a maioria das tarefas."),
            ("Low",       "16A34A", "Baixa prioridade. Pode ser remanejada."),
            ("(vazio)",   "CBD5E1", "Sem prioridade definida."),
        ],
    },
    "es": {
        "lang_full": "es",
        "file_suffix": "es",
        "sheet_data": "Datos",
        "sheet_instructions": "Instrucciones",
        "sheet_status": "Glosario Estado",
        "sheet_completion": "% Progreso",
        "ws2_title": "Cómo completar esta plantilla — Diagrama de Gantt Interactivo",
        "ws2_subtitle": "Complete la hoja 'Datos' e importe este archivo en gantt.html mediante el botón Importar.",
        "header_required": "Obligatorio",
        "header_type": "Tipo",
        "header_column": "Columna",
        "header_desc": "Descripción y reglas",
        "yes": "Sí",
        "no": "No",
        "rules_title": "Reglas generales y consejos",
        "status_ref_title": "Estado (prioridad) — referencia visual",
        "status_when": "Cuándo usar",
        "status_color": "Color",
        "status_value": "Valor",
        "comp_title": "Campo 'completion' — lógica detallada",
        "comp_intro": ("Porcentaje de finalización de la tarea, rango 0 a 100. Esta hoja explica cómo el valor "
                       "es interpretado por el Gantt, cómo las fases agregan, y qué criterio usar al completar."),
        "comp_s1": "1. Cómo lo representa visualmente la app",
        "comp_s1_text": (
            "La barra de la tarea se dibuja en dos capas superpuestas:\n"
            "  • Una barra de fondo (color claro, 30% de opacidad) que representa la duración total.\n"
            "  • Una barra de progreso (color sólido) que ocupa exactamente 'completion'% del ancho de fondo.\n"
            "Es decir: completion=0 → solo fondo claro;  completion=50 → mitad llena;  completion=100 → barra totalmente llena."),
        "comp_visual_title": "Ejemplo visual (ancho de la barra = % completado):",
        "comp_s2": "2. Los hitos (milestone=TRUE) ignoran este campo",
        "comp_s2_text": (
            "Los hitos son puntos en el tiempo, no períodos — por eso la noción de 'finalización parcial' no aplica.\n"
            "La app representa el hito como un rombo lleno, independientemente del valor en 'completion'.\n"
            "Convención: deje completion=0 en los hitos."),
        "comp_who_title": "3. Quién completa y quién calcula la app",
        "comp_who_intro": "Hay 3 tipos de filas. La regla de oro: usted completa las HOJAS, la app calcula los RESÚMENES.",
        "comp_who_headers": ["Tipo de fila", "¿Quién decide el %?", "Ejemplo en el proyecto"],
        "comp_who_rows": [
            ["Hoja (sin hijos)", "Usted la completa manualmente (0-100)", "Wireframes = 80"],
            ["Resumen (tiene hijos)", "App calcula. NO complete — será sobrescrito.", "Design = promedio de hojas"],
            ["Hito (milestone=TRUE)", "Deje 0. La app lo ignora en el agregado.", "Go-Live"],
        ],
        "comp_s3": "4. Fórmula recursiva (vale para cualquier profundidad)",
        "comp_s3_text": (
            "Para cada nodo-resumen (cualquier nivel: proyecto, workstream, fase, sub-fase), la app:\n"
            "  1. Recorre TODO el árbol descendente recogiendo las HOJAS\n"
            "  2. Descarta las que son hitos (milestone=TRUE)\n"
            "  3. Calcula el PROMEDIO SIMPLE de los % de esas hojas\n\n"
            "Pseudocódigo:\n"
            "    función calcular_pct(nodo):\n"
            "        si nodo NO tiene hijos:           ← es hoja\n"
            "            retorna nodo.completion\n"
            "        sino:\n"
            "            hojas = todas las descendientes-hoja (recursión)\n"
            "            válidas = hojas donde milestone == FALSE\n"
            "            si válidas vacías:\n"
            "                retorna 0\n"
            "            retorna SUMA(válidas[i].completion) / CUENTA(válidas)\n\n"
            "Notación matemática:\n"
            "    %_resumen = Σ(%_hojas_no_hito) / N_hojas_no_hito"),
        "comp_ex_title": "5. Ejemplo trabajado completo — 3 niveles, 13 hojas",
        "comp_ex_intro": ("Considere el dataset por defecto (Proyecto Ejemplo). Vea el cálculo bottom-up — "
                          "cada fila muestra nivel, tipo, fórmula aplicada y resultado:"),
        "comp_ex_headers": ["Nodo", "Nivel", "Tipo", "Cálculo", "Resultado"],
        "comp_ex_rows": [
            ["Requirements gathering",    "2", "hoja",   "—", "50"],
            ["Requirements gathering 2",  "2", "hoja",   "—", "55"],
            ["Stakeholder workshop",      "2", "hoja",   "—", "75"],
            ["Story boarding",            "2", "hoja",   "—", "80"],
            ["Initiation complete",       "2", "hito",   "(ignorado)", "—"],
            ["▸ Initiation",              "1", "resumen","(50+55+75+80)/4", "65%"],
            ["E2E data solution design",  "2", "hoja",   "—", "100"],
            ["Wireframes",                "2", "hoja",   "—", "80"],
            ["Prototyping",               "2", "hoja",   "—", "40"],
            ["Design complete",           "2", "hito",   "(ignorado)", "—"],
            ["▸ Design",                  "1", "resumen","(100+80+40)/3", "73%"],
            ["ETL",                       "2", "hoja",   "—", "15"],
            ["Data modelling",            "2", "hoja",   "—", "40"],
            ["Measures & KPIs",           "2", "hoja",   "—", "50"],
            ["Visuals",                   "2", "hoja",   "—", "25"],
            ["Implementation complete",   "2", "hito",   "(ignorado)", "—"],
            ["▸ Implementation",          "1", "resumen","(15+40+50+25)/4", "33%"],
            ["UAT",                       "2", "hoja",   "—", "0"],
            ["Bug fixes",                 "2", "hoja",   "—", "0"],
            ["Go-Live",                   "2", "hito",   "(ignorado)", "—"],
            ["▸ Testing",                 "1", "resumen","(0+0)/2", "0%"],
            ["■ PROYECTO EJEMPLO",        "0", "raíz",   "suma de las 13 hojas / 13 = 610/13", "47%"],
        ],
        "comp_ex_note": ("Verificación de la raíz: 50+55+75+80 + 100+80+40 + 15+40+50+25 + 0+0 = 610. "
                          "610 ÷ 13 = 46,92% → redondeado a 47%."),
        "comp_warning_title": "6. ATENCIÓN: promedio de HOJAS, no de fases",
        "comp_warning_body": ("El % de la raíz NO es el promedio de las fases — es el promedio de las hojas DIRECTAMENTE. "
                              "Las fases con número diferente de tareas pesan distinto en los dos métodos:"),
        "comp_warning_headers": ["Método de cálculo", "Fórmula", "Resultado"],
        "comp_warning_rows": [
            ["Promedio de 4 fases (INCORRECTO)",   "(65 + 73 + 33 + 0) / 4",   "43%"],
            ["Promedio de 13 hojas (CORRECTO)",    "610 / 13",                  "47%"],
        ],
        "comp_warning_explain": ("La diferencia existe porque Testing (solo 2 hojas, ambas 0%) reduce MENOS el resultado "
                                  "cuando miramos hoja por hoja. La app usa promedio de hojas — decisión consciente. "
                                  "Si quiere promedio ponderado por duración o esfuerzo, calcule externamente "
                                  "(el modelo no tiene columna de peso)."),
        "comp_s4": "4. Criterio sugerido (rúbrica)",
        "comp_s4_text": ("Use la rúbrica para consistencia. Trabaje en múltiplos de 25 — "
                          "mayor precisión rara vez refleja realidad observable."),
        "comp_rubric_headers": ["%", "Significado", "Cuándo usar"],
        "comp_rubric_rows": [
            ["0",   "No iniciada",     "La tarea aún no ha comenzado."],
            ["25",  "Iniciada",        "Trabajo empezado. Entregable parcial."],
            ["50",  "A medio camino",  "Mitad del esfuerzo gastado."],
            ["75",  "Casi lista",      "Revisada al menos una vez."],
            ["100", "Completada",      "Entregable aceptado por el responsable."],
        ],
        "comp_s5": "5. Trampas comunes",
        "comp_gotchas": [
            ("No use porcentaje temporal (días transcurridos / duración total).",
             "Ese cálculo no dice nada sobre progreso real del entregable."),
            ("No complete 'completion' en la fila de la fase.",
             "Se deriva automáticamente."),
            ("No use valores fraccionarios (37,5%).",
             "La app acepta solo enteros."),
            ("100% ≠ 'aceptado por el cliente'.",
             "Aceptación formal normalmente es un hito aparte."),
            ("Cuidado al volver tarea en hito con completion > 0.",
             "El valor será ignorado; póngalo en 0 manualmente."),
            ("100% con 'end' en el futuro es sospechoso.",
             "Actualice 'end' a la fecha real de término."),
        ],
        "comp_summary_label": "6. Resumen de 1 línea",
        "comp_summary": ("'completion' es PORCENTAJE DE ENTREGABLE producido (no tiempo transcurrido), completado "
                          "por el equipo en múltiplos de 25, ignorado en hitos, agregado por promedio simple en la fila de fase."),
        "columns": [
            ("id", 8, "Identificador único de la tarea. Entero secuencial. Obligatorio para dependencias."),
            ("parent_id", 12, "ID de la tarea padre. Vacío = raíz. Define jerarquía N niveles (proyecto > fase > tarea > sub-tarea...)."),
            ("phase", 22, "[Legado] Nombre de fase. Solo si parent_id está vacío. Recomendado: dejar vacío y usar parent_id."),
            ("task", 38, "Descripción de la tarea que aparece en la barra del Gantt."),
            ("milestone", 12, "TRUE = hito (rombo), FALSE = barra normal."),
            ("start", 14, "Fecha de inicio. Formato DD/MM/AAAA. En tareas-resumen, dejar vacío (calculado de hijos)."),
            ("end", 14, "Fecha de término. Formato DD/MM/AAAA. En tareas-resumen, dejar vacío."),
            ("completion", 12, "Porcentaje completado (0 a 100). En resumen, calculado de hijos. Ver hoja '% Progreso'."),
            ("dependencies", 16, "IDs predecesores separados por coma. Ej: 2,3"),
            ("assignee", 18, "Responsable de la tarea (texto libre)."),
            ("status", 14, "Prioridad: Very High | High | Medium | Low | (vacío)."),
        ],
        "details": [
            ("id", "Entero", "Sí", "Identificador único. Use números secuenciales. Referenciado por 'dependencies' y 'parent_id'."),
            ("parent_id", "Entero", "No", "ID de la tarea padre. Vacío = nodo raíz. Define jerarquía: Proyecto > Workstream > Fase > Sub-fase > Tarea. Profundidad ilimitada. Los ciclos se detectan y bloquean."),
            ("phase", "Texto", "No", "[Legado] Si parent_id está vacío pero phase tiene valor, el app crea una fase-resumen automáticamente. Recomendado: dejar vacío y usar parent_id explícitamente."),
            ("task", "Texto", "Sí", "Descripción que aparece en la barra. Idealmente hasta 60 caracteres."),
            ("milestone", "TRUE/FALSE", "No", "TRUE = rombo. FALSE = barra. Para hitos, start = end."),
            ("start", "Fecha DD/MM/AAAA", "Sí", "Fecha de inicio. Acepta formato fecha nativo de Excel."),
            ("end", "Fecha DD/MM/AAAA", "Sí", "Fecha de término. Debe ser ≥ start."),
            ("completion", "Entero 0-100", "No", "Porcentaje de entregable producido. Ver hoja '% Progreso'."),
            ("dependencies", "Texto", "No", "IDs de predecesoras separados por coma. Ej.: '2,3'."),
            ("assignee", "Texto", "No", "Responsable (texto libre). Aparece en el tooltip."),
            ("status", "Lista", "No", "Very High / High / Medium / Low / (vacío). Define color de la barra."),
        ],
        "rules": [
            "1. No deje filas en blanco entre tareas — el importador para en la primera vacía.",
            "2. Los IDs no necesitan ser secuenciales sin huecos, pero sí únicos y estables.",
            "3. Fechas: acepta fecha Excel y texto DD/MM/AAAA.",
            "4. Para crear una fase nueva, escriba un nombre nuevo en 'phase'.",
            "5. Los hitos representan puntos en el tiempo — siempre start = end.",
            "6. Dependencias circulares generan flechas inconsistentes. Evítelas.",
            "7. Estado en blanco = gris neutro (útil para hitos genéricos).",
            "8. Guarde, abra gantt.html, haga clic en Importar, elija 'Reemplazar todo'.",
            "9. Puede mezclar este Excel con edición inline después.",
            "10. Para colaboración por Git, exporte como CSV.",
        ],
        "status_ref": [
            ("Very High", "B91C1C", "Crítico. Camino crítico o bloqueador absoluto."),
            ("High",      "EA580C", "Alta prioridad. Retrasos tienen impacto significativo."),
            ("Medium",    "CA8A04", "Estándar para la mayoría de tareas."),
            ("Low",       "16A34A", "Baja prioridad. Puede reorganizarse."),
            ("(vacío)",   "CBD5E1", "Sin prioridad definida."),
        ],
    },
    "en": {
        "lang_full": "en",
        "file_suffix": "en",
        "sheet_data": "Data",
        "sheet_instructions": "Instructions",
        "sheet_status": "Status Glossary",
        "sheet_completion": "% Completion",
        "ws2_title": "How to fill this template — Interactive Gantt Chart",
        "ws2_subtitle": "Fill the 'Data' sheet and import this file into gantt.html via the Import button.",
        "header_required": "Required",
        "header_type": "Type",
        "header_column": "Column",
        "header_desc": "Description and rules",
        "yes": "Yes",
        "no": "No",
        "rules_title": "General rules and tips",
        "status_ref_title": "Status (priority) — visual reference",
        "status_when": "When to use",
        "status_color": "Color",
        "status_value": "Value",
        "comp_title": "'completion' field — detailed logic",
        "comp_intro": ("Task completion percentage, range 0 to 100. This sheet explains how the value is "
                       "interpreted by the Gantt, how phases aggregate, and what criteria to use when filling it."),
        "comp_s1": "1. How the app renders it visually",
        "comp_s1_text": (
            "The task bar is drawn in two overlapping layers:\n"
            "  • A background bar (light color, 30% opacity) representing the total duration.\n"
            "  • A progress bar (solid color) occupying exactly 'completion'% of the background width.\n"
            "So: completion=0 → only light background;  completion=50 → half filled;  completion=100 → fully filled."),
        "comp_visual_title": "Visual example (bar width = % filled):",
        "comp_s2": "2. Milestones (milestone=TRUE) ignore this field",
        "comp_s2_text": (
            "Milestones are points in time, not periods — so the concept of 'partial completion' doesn't apply.\n"
            "The app renders milestones as filled diamonds, regardless of the 'completion' value.\n"
            "Convention: leave completion=0 on milestones."),
        "comp_who_title": "3. Who fills and who the app computes",
        "comp_who_intro": "There are 3 row types. Golden rule: you fill LEAVES, the app computes SUMMARIES.",
        "comp_who_headers": ["Row type", "Who decides the %?", "Example in default project"],
        "comp_who_rows": [
            ["Leaf (no children)", "You fill it manually (0-100)", "Wireframes = 80"],
            ["Summary (has children)", "App computes. DON'T fill — it'll be overwritten.", "Design = avg of leaves"],
            ["Milestone (milestone=TRUE)", "Leave 0. App ignores in aggregation.", "Go-Live"],
        ],
        "comp_s3": "4. Recursive formula (works at any depth)",
        "comp_s3_text": (
            "For each summary node (any level: project, workstream, phase, sub-phase), the app:\n"
            "  1. Walks the ENTIRE descendant tree collecting LEAVES\n"
            "  2. Discards milestones (milestone=TRUE)\n"
            "  3. Computes SIMPLE AVERAGE of those leaves' %\n\n"
            "Pseudocode:\n"
            "    function compute_pct(node):\n"
            "        if node has NO children:        ← it's a leaf\n"
            "            return node.completion\n"
            "        else:\n"
            "            leaves = all descendant leaves (recursion)\n"
            "            valid = leaves where milestone == FALSE\n"
            "            if valid is empty:\n"
            "                return 0\n"
            "            return SUM(valid[i].completion) / COUNT(valid)\n\n"
            "Mathematical notation:\n"
            "    %_summary = Σ(%_non_milestone_leaves) / N_non_milestone_leaves"),
        "comp_ex_title": "5. Full worked example — 3 levels, 13 leaves",
        "comp_ex_intro": ("Consider the default dataset (Sample Project). See the bottom-up calculation — "
                          "each row shows level, type, applied formula and result:"),
        "comp_ex_headers": ["Node", "Level", "Type", "Calculation", "Result"],
        "comp_ex_rows": [
            ["Requirements gathering",    "2", "leaf",     "—", "50"],
            ["Requirements gathering 2",  "2", "leaf",     "—", "55"],
            ["Stakeholder workshop",      "2", "leaf",     "—", "75"],
            ["Story boarding",            "2", "leaf",     "—", "80"],
            ["Initiation complete",       "2", "milestone","(ignored)", "—"],
            ["▸ Initiation",              "1", "summary",  "(50+55+75+80)/4", "65%"],
            ["E2E data solution design",  "2", "leaf",     "—", "100"],
            ["Wireframes",                "2", "leaf",     "—", "80"],
            ["Prototyping",               "2", "leaf",     "—", "40"],
            ["Design complete",           "2", "milestone","(ignored)", "—"],
            ["▸ Design",                  "1", "summary",  "(100+80+40)/3", "73%"],
            ["ETL",                       "2", "leaf",     "—", "15"],
            ["Data modelling",            "2", "leaf",     "—", "40"],
            ["Measures & KPIs",           "2", "leaf",     "—", "50"],
            ["Visuals",                   "2", "leaf",     "—", "25"],
            ["Implementation complete",   "2", "milestone","(ignored)", "—"],
            ["▸ Implementation",          "1", "summary",  "(15+40+50+25)/4", "33%"],
            ["UAT",                       "2", "leaf",     "—", "0"],
            ["Bug fixes",                 "2", "leaf",     "—", "0"],
            ["Go-Live",                   "2", "milestone","(ignored)", "—"],
            ["▸ Testing",                 "1", "summary",  "(0+0)/2", "0%"],
            ["■ SAMPLE PROJECT",          "0", "root",     "sum of 13 leaves / 13 = 610/13", "47%"],
        ],
        "comp_ex_note": ("Root verification: 50+55+75+80 + 100+80+40 + 15+40+50+25 + 0+0 = 610. "
                          "610 ÷ 13 = 46.92% → rounded to 47%."),
        "comp_warning_title": "6. CAUTION: average of LEAVES, not of phases",
        "comp_warning_body": ("The root's % is NOT the average of phases — it's the average of leaves DIRECTLY. "
                              "Phases with different task counts weigh differently between the two methods:"),
        "comp_warning_headers": ["Calculation method", "Formula", "Result"],
        "comp_warning_rows": [
            ["Average of 4 phases (WRONG)",     "(65 + 73 + 33 + 0) / 4",   "43%"],
            ["Average of 13 leaves (CORRECT)",  "610 / 13",                  "47%"],
        ],
        "comp_warning_explain": ("The difference exists because Testing (only 2 leaves, both 0%) pulls the result down LESS "
                                  "when we look leaf-by-leaf. The app uses leaf average — deliberate choice. "
                                  "For weighted average by duration or effort, compute externally "
                                  "(the model has no weight column)."),
        "comp_s4": "4. Suggested criteria (rubric)",
        "comp_s4_text": ("Use the rubric below for consistency. Work in multiples of 25 — "
                          "finer precision rarely reflects observable reality."),
        "comp_rubric_headers": ["%", "Meaning", "When to use"],
        "comp_rubric_rows": [
            ["0",   "Not started",   "Task has not begun."],
            ["25",  "Started",       "Work began. Partial deliverable exists."],
            ["50",  "Halfway",       "About half of the effort spent."],
            ["75",  "Almost done",   "Reviewed at least once. Final tweaks."],
            ["100", "Completed",     "Deliverable accepted by owner."],
        ],
        "comp_s5": "5. Common gotchas",
        "comp_gotchas": [
            ("Don't use temporal percentage (days elapsed / total duration).",
             "That calculation says nothing about actual deliverable progress."),
            ("Don't fill 'completion' on the phase row.",
             "It's auto-derived from child tasks."),
            ("Don't use fractional values (37.5%).",
             "The app accepts integers only."),
            ("100% ≠ 'client-approved'.",
             "Formal approval is usually a separate milestone."),
            ("Be careful flipping a task to milestone with completion > 0.",
             "The value will be ignored; zero it out manually for hygiene."),
            ("100% with 'end' in the future is suspicious.",
             "Update 'end' to the actual completion date."),
        ],
        "comp_summary_label": "6. One-line summary",
        "comp_summary": ("'completion' is PERCENT OF DELIVERABLE produced (not time elapsed), filled by "
                          "the task team in multiples of 25, ignored on milestones, aggregated by simple average on the phase row."),
        "columns": [
            ("id", 8, "Unique task identifier. Sequential integer. Required for dependency references."),
            ("parent_id", 12, "ID of parent task. Empty = root. Defines N-level hierarchy (project > phase > task > sub-task...)."),
            ("phase", 22, "[Legacy] Phase name. Only used if parent_id is empty. Recommended: leave empty and use parent_id."),
            ("task", 38, "Task description shown on the Gantt bar."),
            ("milestone", 12, "TRUE = milestone (diamond), FALSE = normal bar."),
            ("start", 14, "Start date. Format DD/MM/YYYY. For summary tasks, leave empty (computed from children)."),
            ("end", 14, "End date. Format DD/MM/YYYY. For summary tasks, leave empty."),
            ("completion", 12, "Percent complete (0-100). For summary, computed from children. See '% Completion' sheet."),
            ("dependencies", 16, "Predecessor IDs comma-separated. Ex: 2,3"),
            ("assignee", 18, "Task owner (free text)."),
            ("status", 14, "Priority: Very High | High | Medium | Low | (empty)."),
        ],
        "details": [
            ("id", "Integer", "Yes", "Unique identifier. Use sequential numbers. Referenced by 'dependencies' and 'parent_id'."),
            ("parent_id", "Integer", "No", "ID of parent task. Empty = root node. Defines hierarchy: Project > Workstream > Phase > Sub-phase > Task. Unlimited depth. Cycles are detected and blocked."),
            ("phase", "Text", "No", "[Legacy] If parent_id is empty but phase is filled, the app auto-creates a phase summary. Recommended: leave empty and use parent_id explicitly."),
            ("task", "Text", "Yes", "Description shown on the bar. Ideally up to 60 characters."),
            ("milestone", "TRUE/FALSE", "No", "TRUE = diamond. FALSE = bar. For milestones, start = end."),
            ("start", "Date DD/MM/YYYY", "Yes", "Start date. Accepts Excel native date format."),
            ("end", "Date DD/MM/YYYY", "Yes", "End date. Must be ≥ start."),
            ("completion", "Integer 0-100", "No", "Percent of deliverable produced. See '% Completion' sheet."),
            ("dependencies", "Text", "No", "Predecessor IDs comma-separated. Ex.: '2,3'."),
            ("assignee", "Text", "No", "Owner (free text). Shown in tooltip."),
            ("status", "List", "No", "Very High / High / Medium / Low / (empty). Sets bar color."),
        ],
        "rules": [
            "1. Don't leave blank rows between tasks — importer stops at the first empty.",
            "2. IDs don't need to be sequential without gaps, but must be unique and stable.",
            "3. Dates: accepts both Excel date and DD/MM/YYYY text.",
            "4. To create a new phase, just type a new name in 'phase'.",
            "5. Milestones represent points in time — always start = end.",
            "6. Circular dependencies produce inconsistent arrows. Avoid them.",
            "7. Empty status = neutral gray (useful for generic milestones).",
            "8. Save, open gantt.html, click Import, choose 'Replace all'.",
            "9. You can mix this Excel with inline editing afterwards.",
            "10. For Git collaboration, export as CSV.",
        ],
        "status_ref": [
            ("Very High", "B91C1C", "Critical. Critical path or absolute blocker."),
            ("High",      "EA580C", "High priority. Delays have significant impact."),
            ("Medium",    "CA8A04", "Default for most tasks."),
            ("Low",       "16A34A", "Low priority. Can be rescheduled."),
            ("(empty)",   "CBD5E1", "No priority defined."),
        ],
    },
}

# ============================================================
# Dados de exemplo (compartilhados nos 3 idiomas — sao identificadores)
# ============================================================
# Estrutura: (id, parent_id, phase, task, milestone, start, end, completion, dependencies, assignee, status)
EXAMPLES = [
    # ROOT — Projeto
    (100, "",    "", "Projeto Exemplo",                  False, "", "", 0,  "",     "",          ""),
    # FASES
    (101, "100", "", "Initiation",                       False, "", "", 0,  "",     "",          ""),
    (102, "100", "", "Design",                           False, "", "", 0,  "",     "",          ""),
    (103, "100", "", "Implementation",                   False, "", "", 0,  "",     "",          ""),
    (104, "100", "", "Testing",                          False, "", "", 0,  "",     "",          ""),
    # TASKS — folhas
    (2,  "101", "", "Requirements gathering",            False, "25/09/2025", "27/09/2025", 50, "",     "",          "High"),
    (3,  "101", "", "Requirements gathering 2",          False, "27/09/2025", "29/09/2025", 55, "2",    "Day Bacci", "Medium"),
    (4,  "101", "", "Stakeholder workshop excl. Execs",  False, "29/09/2025", "01/10/2025", 75, "3",    "",          "Medium"),
    (5,  "101", "", "Story boarding",                    False, "29/09/2025", "07/10/2025", 80, "3",    "",          "Low"),
    (6,  "101", "", "Initiation complete",               True,  "07/10/2025", "07/10/2025", 0,  "4,5",  "",          ""),
    (8,  "102", "", "E2E data solution design",          False, "01/10/2025", "03/10/2025", 100,"6",    "",          "High"),
    (9,  "102", "", "Wireframes",                        False, "03/10/2025", "08/10/2025", 80, "8",    "",          "Medium"),
    (10, "102", "", "Prototyping",                       False, "07/10/2025", "17/10/2025", 40, "9",    "",          "Low"),
    (11, "102", "", "Design complete",                   True,  "17/10/2025", "17/10/2025", 0,  "10",   "",          ""),
    (13, "103", "", "ETL",                               False, "03/10/2025", "14/10/2025", 15, "6",    "",          "Very High"),
    (14, "103", "", "Data modelling",                    False, "07/10/2025", "14/10/2025", 40, "13",   "",          "High"),
    (15, "103", "", "Measures & KPIs",                   False, "07/10/2025", "10/10/2025", 50, "14",   "",          "Medium"),
    (16, "103", "", "Visuals",                           False, "10/10/2025", "20/10/2025", 25, "15",   "",          "Medium"),
    (17, "103", "", "Implementation complete",           True,  "24/10/2025", "24/10/2025", 0,  "16",   "",          ""),
    (18, "104", "", "UAT",                               False, "20/10/2025", "29/10/2025", 0,  "16",   "",          "High"),
    (19, "104", "", "Bug fixes",                         False, "24/10/2025", "30/10/2025", 0,  "17",   "",          "Medium"),
    (20, "104", "", "Go-Live",                           True,  "31/10/2025", "31/10/2025", 0,  "18,19","",          "Very High"),
]


def section_header(ws, row, title):
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 24


def build_template(lang_key):
    L = TR[lang_key]
    wb = Workbook()

    # ---------- ABA 1: Data ----------
    ws = wb.active
    ws.title = L["sheet_data"]

    columns = L["columns"]
    for col_idx, (name, width, _desc) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for row_idx, row_data in enumerate(EXAMPLES, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGN
            cell.border = BORDER
            # column 11 = status (após adicionar parent_id)
            if col_idx == 11:
                colors = {"Very High":"FECACA", "High":"FED7AA", "Medium":"FEF08A", "Low":"BBF7D0"}
                if value in colors:
                    cell.fill = PatternFill("solid", fgColor=colors[value])
            # column 5 = milestone
            if col_idx == 5 and value:
                cell.fill = PatternFill("solid", fgColor="FEF3C7")
                cell.font = Font(name="Calibri", size=11, bold=True)
            # centered: id(1), parent_id(2), milestone(5), completion(8)
            if col_idx in (1, 2, 5, 8):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # destaque visual em linhas de resumo (parent_id vazio = raiz; ou tarefa sem datas = resumo)
            if col_idx == 4 and row_data[1] == "" and not row_data[5]:
                cell.font = Font(name="Calibri", size=11, bold=True, color="1F2937")
                cell.fill = PatternFill("solid", fgColor="EEF2F7")

    LAST_ROW = 500
    dv_milestone = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    dv_milestone.add(f"E2:E{LAST_ROW}")
    ws.add_data_validation(dv_milestone)
    dv_status = DataValidation(type="list", formula1='"Very High,High,Medium,Low"', allow_blank=True)
    dv_status.add(f"K2:K{LAST_ROW}")
    ws.add_data_validation(dv_status)
    dv_completion = DataValidation(type="whole", operator="between", formula1=0, formula2=100, allow_blank=True)
    dv_completion.add(f"H2:H{LAST_ROW}")
    ws.add_data_validation(dv_completion)

    for r in range(len(EXAMPLES) + 2, 51):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.font = CELL_FONT

    # ---------- ABA 2: Instructions ----------
    ws2 = wb.create_sheet(L["sheet_instructions"])
    ws2["A1"] = L["ws2_title"]
    ws2["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F2937")
    ws2.merge_cells("A1:D1")
    ws2.row_dimensions[1].height = 28
    ws2["A2"] = L["ws2_subtitle"]
    ws2["A2"].font = Font(name="Calibri", size=11, italic=True, color="6B7280")
    ws2.merge_cells("A2:D2")

    ws2["A4"] = L["header_column"]
    ws2["B4"] = L["header_type"]
    ws2["C4"] = L["header_required"]
    ws2["D4"] = L["header_desc"]
    for col in ["A4", "B4", "C4", "D4"]:
        ws2[col].font = HEADER_FONT
        ws2[col].fill = HEADER_FILL
        ws2[col].alignment = HEADER_ALIGN
        ws2[col].border = BORDER
    ws2.row_dimensions[4].height = 24

    for i, (col, tipo, obrig, desc) in enumerate(L["details"], start=5):
        ws2.cell(row=i, column=1, value=col).font = Font(name="Consolas", size=11, bold=True, color="1E40AF")
        ws2.cell(row=i, column=2, value=tipo).font = CELL_FONT
        obrig_label = L["yes"] if obrig in ("Sim", "Sí", "Yes") else L["no"]
        obrig_cell = ws2.cell(row=i, column=3, value=obrig_label)
        obrig_cell.font = Font(name="Calibri", size=11, bold=True,
                                color="DC2626" if obrig_label == L["yes"] else "6B7280")
        obrig_cell.alignment = Alignment(horizontal="center", vertical="center")
        desc_cell = ws2.cell(row=i, column=4, value=desc)
        desc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        desc_cell.font = CELL_FONT
        for c in range(1, 5):
            ws2.cell(row=i, column=c).border = BORDER
            ws2.cell(row=i, column=c).alignment = Alignment(
                horizontal=ws2.cell(row=i, column=c).alignment.horizontal or "left",
                vertical="top", wrap_text=True,
            )
        ws2.row_dimensions[i].height = 60

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 90

    start_rules = len(L["details"]) + 7
    ws2.cell(row=start_rules, column=1, value=L["rules_title"]).font = Font(name="Calibri", size=14, bold=True, color="1F2937")
    ws2.merge_cells(start_row=start_rules, start_column=1, end_row=start_rules, end_column=4)
    for i, rule in enumerate(L["rules"], start=start_rules + 1):
        cell = ws2.cell(row=i, column=1, value=rule)
        cell.font = CELL_FONT
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws2.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
        ws2.row_dimensions[i].height = 22

    # ---------- ABA 3: Status Glossary ----------
    ws3 = wb.create_sheet(L["sheet_status"])
    ws3["A1"] = L["status_ref_title"]
    ws3["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws3.merge_cells("A1:C1")
    ws3["A3"] = L["status_value"]
    ws3["B3"] = L["status_color"]
    ws3["C3"] = L["status_when"]
    for col in ["A3", "B3", "C3"]:
        ws3[col].font = HEADER_FONT
        ws3[col].fill = HEADER_FILL
        ws3[col].alignment = HEADER_ALIGN
        ws3[col].border = BORDER
    for i, (val, color, desc) in enumerate(L["status_ref"], start=4):
        ws3.cell(row=i, column=1, value=val).font = Font(name="Calibri", size=11, bold=True)
        color_cell = ws3.cell(row=i, column=2, value="")
        color_cell.fill = PatternFill("solid", fgColor=color)
        ws3.cell(row=i, column=3, value=desc).font = CELL_FONT
        for c in range(1, 4):
            ws3.cell(row=i, column=c).border = BORDER
            ws3.cell(row=i, column=c).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws3.row_dimensions[i].height = 26
    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 70

    # ---------- ABA 4: % Completion (didática) ----------
    ws4 = wb.create_sheet(L["sheet_completion"])

    # Larguras: A bem larga, depois 20 colunas estreitas para as barras visuais,
    # depois colunas mais largas para a tabela do exemplo
    ws4.column_dimensions["A"].width = 36
    for c in range(2, 22):
        ws4.column_dimensions[get_column_letter(c)].width = 3
    ws4.column_dimensions["W"].width = 12   # Nível
    ws4.column_dimensions["X"].width = 12   # Tipo
    ws4.column_dimensions["Y"].width = 38   # Cálculo
    ws4.column_dimensions["Z"].width = 14   # Resultado

    # Title + intro
    ws4["A1"] = L["comp_title"]
    ws4["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F2937")
    ws4.merge_cells("A1:Z1")
    ws4.row_dimensions[1].height = 28
    ws4["A2"] = L["comp_intro"]
    ws4["A2"].font = Font(name="Calibri", size=11, italic=True, color="6B7280")
    ws4["A2"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws4.merge_cells("A2:Z2")
    ws4.row_dimensions[2].height = 40

    # SECTION 1: Visual rendering (mantém na mesma área visual das barras)
    section_header(ws4, 4, L["comp_s1"])
    ws4["A5"] = L["comp_s1_text"]
    ws4["A5"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws4["A5"].font = CELL_FONT
    ws4.merge_cells("A5:Z5")
    ws4.row_dimensions[5].height = 90

    ws4["A7"] = L["comp_visual_title"]
    ws4["A7"].font = Font(name="Calibri", size=11, bold=True)
    ws4.merge_cells("A7:Z7")

    visual_rows = [(8, "0%", 0), (9, "25%", 25), (10, "50%", 50), (11, "75%", 75), (12, "100%", 100)]
    for row, label, pct in visual_rows:
        ws4.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=11, bold=True)
        ws4.cell(row=row, column=1).alignment = Alignment(horizontal="right", vertical="center")
        cells_filled = round(pct / 5)
        for c in range(2, 22):
            cell = ws4.cell(row=row, column=c, value=None)
            if c - 1 <= cells_filled:
                cell.fill = PatternFill("solid", fgColor="2563EB")
            else:
                cell.fill = PatternFill("solid", fgColor="DBEAFE")
            cell.border = Border(
                left=Side(style="thin", color="2563EB"),
                right=Side(style="thin", color="2563EB"),
                top=Side(style="thin", color="2563EB"),
                bottom=Side(style="thin", color="2563EB"),
            )
        ws4.row_dimensions[row].height = 20

    # SECTION 2: Marcos
    section_header(ws4, 14, L["comp_s2"])
    ws4["A15"] = L["comp_s2_text"]
    ws4["A15"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws4["A15"].font = CELL_FONT
    ws4.merge_cells("A15:Z15")
    ws4.row_dimensions[15].height = 70

    # SECTION 3: Quem preenche (tabela de decisão)
    section_header(ws4, 17, L["comp_who_title"])
    ws4["A18"] = L["comp_who_intro"]
    ws4["A18"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws4["A18"].font = CELL_FONT
    ws4.merge_cells("A18:Z18")
    ws4.row_dimensions[18].height = 30

    # Cabeçalhos da tabela (cols A, W, Y) - mas usar A, W, Y, Z não rende bem por causa do span
    # Mais simples: usar cols A (nome), W (decide), Y (exemplo) - mesclar a coluna decide
    who_target_cols = [1, 23, 25]  # A, W, Y
    for col_idx, v in zip(who_target_cols, L["comp_who_headers"]):
        cell = ws4.cell(row=19, column=col_idx, value=v)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
        # estender col X e Z para o lado direito do header
    # mesclar cabeçalho W:X e Y:Z
    ws4.merge_cells(start_row=19, start_column=23, end_row=19, end_column=24)
    ws4.merge_cells(start_row=19, start_column=25, end_row=19, end_column=26)
    ws4.row_dimensions[19].height = 24

    for r_idx, row_data in enumerate(L["comp_who_rows"], start=20):
        ws4.cell(row=r_idx, column=1, value=row_data[0]).font = Font(name="Calibri", size=11, bold=True)
        ws4.cell(row=r_idx, column=23, value=row_data[1]).font = CELL_FONT
        ws4.merge_cells(start_row=r_idx, start_column=23, end_row=r_idx, end_column=24)
        ws4.cell(row=r_idx, column=25, value=row_data[2]).font = Font(name="Consolas", size=10, color="6B7280")
        ws4.merge_cells(start_row=r_idx, start_column=25, end_row=r_idx, end_column=26)
        for c in (1, 23, 25):
            ws4.cell(row=r_idx, column=c).border = BORDER
            ws4.cell(row=r_idx, column=c).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws4.row_dimensions[r_idx].height = 30

    # SECTION 4: Fórmula recursiva
    section_header(ws4, 24, L["comp_s3"])
    ws4["A25"] = L["comp_s3_text"]
    ws4["A25"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws4["A25"].font = Font(name="Consolas", size=10)
    ws4.merge_cells("A25:Z25")
    ws4.row_dimensions[25].height = 280

    # SECTION 5: Exemplo trabalhado 3 níveis
    section_header(ws4, 27, L["comp_ex_title"])
    ws4["A28"] = L["comp_ex_intro"]
    ws4["A28"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws4["A28"].font = CELL_FONT
    ws4.merge_cells("A28:Z28")
    ws4.row_dimensions[28].height = 40

    # Cabeçalhos da tabela do exemplo (cols A, W, X, Y, Z)
    ex_target_cols = [1, 23, 24, 25, 26]
    for col_idx, v in zip(ex_target_cols, L["comp_ex_headers"]):
        cell = ws4.cell(row=29, column=col_idx, value=v)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws4.row_dimensions[29].height = 22

    for r_idx, row_data in enumerate(L["comp_ex_rows"], start=30):
        is_summary = row_data[2] in ("resumo", "resumen", "summary")
        is_root = row_data[2] in ("raiz", "raíz", "root")
        for col_idx, v in zip(ex_target_cols, row_data):
            cell = ws4.cell(row=r_idx, column=col_idx, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font = CELL_FONT
            if is_summary:
                cell.fill = PatternFill("solid", fgColor="F1F5F9")
                cell.font = Font(name="Calibri", size=11, bold=True)
            elif is_root:
                cell.fill = PatternFill("solid", fgColor="DBEAFE")
                cell.font = Font(name="Calibri", size=11, bold=True, color="1E40AF")

    last_ex_row = 29 + len(L["comp_ex_rows"])
    note_row = last_ex_row + 2
    ws4.cell(row=note_row, column=1, value=L["comp_ex_note"]).font = Font(name="Consolas", size=10, italic=True, color="6B7280")
    ws4.cell(row=note_row, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws4.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=26)
    ws4.row_dimensions[note_row].height = 50

    # SECTION 6: Atenção média de folhas
    warn_section_row = note_row + 2
    section_header(ws4, warn_section_row, L["comp_warning_title"])
    ws4.cell(row=warn_section_row + 1, column=1, value=L["comp_warning_body"]).font = CELL_FONT
    ws4.cell(row=warn_section_row + 1, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws4.merge_cells(start_row=warn_section_row + 1, start_column=1, end_row=warn_section_row + 1, end_column=26)
    ws4.row_dimensions[warn_section_row + 1].height = 50

    warn_header_row = warn_section_row + 3
    warn_target_cols = [1, 23, 25]
    for col_idx, v in zip(warn_target_cols, L["comp_warning_headers"]):
        cell = ws4.cell(row=warn_header_row, column=col_idx, value=v)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws4.merge_cells(start_row=warn_header_row, start_column=23, end_row=warn_header_row, end_column=24)
    ws4.merge_cells(start_row=warn_header_row, start_column=25, end_row=warn_header_row, end_column=26)

    for offset, row_data in enumerate(L["comp_warning_rows"], start=1):
        rr = warn_header_row + offset
        is_wrong = ("ERRADO" in row_data[0]) or ("INCORRECTO" in row_data[0]) or ("WRONG" in row_data[0])
        fill = PatternFill("solid", fgColor="FEE2E2") if is_wrong else PatternFill("solid", fgColor="DCFCE7")
        ws4.cell(row=rr, column=1, value=row_data[0]).font = Font(name="Calibri", size=11, bold=True)
        ws4.cell(row=rr, column=23, value=row_data[1]).font = Font(name="Consolas", size=10)
        ws4.merge_cells(start_row=rr, start_column=23, end_row=rr, end_column=24)
        ws4.cell(row=rr, column=25, value=row_data[2]).font = Font(name="Calibri", size=12, bold=True)
        ws4.merge_cells(start_row=rr, start_column=25, end_row=rr, end_column=26)
        for c in (1, 23, 25):
            ws4.cell(row=rr, column=c).border = BORDER
            ws4.cell(row=rr, column=c).fill = fill
            ws4.cell(row=rr, column=c).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws4.row_dimensions[rr].height = 28

    warn_explain_row = warn_header_row + len(L["comp_warning_rows"]) + 2
    ws4.cell(row=warn_explain_row, column=1, value=L["comp_warning_explain"]).font = Font(name="Calibri", size=11, italic=True)
    ws4.cell(row=warn_explain_row, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws4.merge_cells(start_row=warn_explain_row, start_column=1, end_row=warn_explain_row, end_column=26)
    ws4.row_dimensions[warn_explain_row].height = 60

    # SECTION 7: Rubrica
    rubric_section_row = warn_explain_row + 2
    section_header(ws4, rubric_section_row, L["comp_s4"])
    ws4.cell(row=rubric_section_row + 1, column=1, value=L["comp_s4_text"]).font = CELL_FONT
    ws4.cell(row=rubric_section_row + 1, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws4.merge_cells(start_row=rubric_section_row + 1, start_column=1, end_row=rubric_section_row + 1, end_column=26)
    ws4.row_dimensions[rubric_section_row + 1].height = 50

    rubric_header_row = rubric_section_row + 3
    for col_idx, v in enumerate(L["comp_rubric_headers"], start=1):
        cell = ws4.cell(row=rubric_header_row, column=col_idx, value=v)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    for offset, row_data in enumerate(L["comp_rubric_rows"], start=1):
        rr = rubric_header_row + offset
        for col_idx, v in enumerate(row_data, start=1):
            cell = ws4.cell(row=rr, column=col_idx, value=v)
            cell.border = BORDER
            cell.font = CELL_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws4.row_dimensions[rr].height = 32

    # SECTION 8: Pegadinhas
    gotchas_section_row = rubric_header_row + len(L["comp_rubric_rows"]) + 2
    section_header(ws4, gotchas_section_row, L["comp_s5"])
    row = gotchas_section_row + 1
    for title, desc in L["comp_gotchas"]:
        cell_t = ws4.cell(row=row, column=1, value="! " + title)
        cell_t.font = Font(name="Calibri", size=11, bold=True, color="B91C1C")
        cell_t.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=26)
        ws4.row_dimensions[row].height = 20
        row += 1
        cell_d = ws4.cell(row=row, column=1, value=desc)
        cell_d.font = CELL_FONT
        cell_d.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=26)
        ws4.row_dimensions[row].height = 36
        row += 1

    # SECTION 9: Resumo
    section_header(ws4, row + 1, L["comp_summary_label"])
    row += 2
    summary = ws4.cell(row=row, column=1, value=L["comp_summary"])
    summary.font = Font(name="Calibri", size=12, italic=True, bold=True, color="1E40AF")
    summary.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    summary.fill = PatternFill("solid", fgColor="EFF6FF")
    summary.border = BORDER
    ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=26)
    ws4.row_dimensions[row].height = 60

    # ---------- Salvar ----------
    out = os.path.join(OUT_DIR, f"gantt-template-{L['file_suffix']}.xlsx")
    wb.save(out)
    print(f"OK: {out}")


for k in ("pt", "es", "en"):
    build_template(k)
