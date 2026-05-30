"""Testes para backend/excel_reader.py"""
import pytest
import openpyxl
from pathlib import Path
import backend.excel_reader as reader_module
from backend.excel_reader import read_excel, create_template


@pytest.fixture(autouse=True)
def reset_cache():
    """Garante que o cache global não vaze entre testes."""
    reader_module._cache = None
    yield
    reader_module._cache = None


def test_read_nonexistent_file():
    data, error = read_excel("nao_existe_jamais.xlsx")
    assert data is None
    assert error is not None
    assert "não encontrado" in error.lower()


def test_create_template_creates_file(tmp_path):
    p = tmp_path / "test.xlsx"
    create_template(str(p))
    assert p.exists()


def test_template_has_all_required_sheets(tmp_path):
    p = tmp_path / "test.xlsx"
    create_template(str(p))
    wb = openpyxl.load_workbook(str(p))
    required = {
        "CONFIG", "FASES", "KPIS", "RESUMO_EXECUTIVO",
        "PENDENCIAS_CRITICAS", "PROXIMAS_ACOES", "CURVA_S", "MARCOS", "RODAPE",
    }
    assert required.issubset(set(wb.sheetnames))
    wb.close()


def test_read_template_returns_all_keys(tmp_path):
    p = tmp_path / "test.xlsx"
    create_template(str(p))
    data, error = read_excel(str(p))
    assert error is None
    assert data is not None
    for key in ("config", "fases", "kpis", "resumo_executivo", "pendencias_criticas",
                "proximas_acoes", "curva_s", "marcos", "rodape"):
        assert key in data, f"Chave '{key}' ausente nos dados lidos"


def test_template_report_date_not_hardcoded(tmp_path):
    """report_date deve ser a data de hoje, não uma data fixa."""
    from datetime import date
    p = tmp_path / "test.xlsx"
    create_template(str(p))
    data, _ = read_excel(str(p))
    today = date.today().strftime("%d/%m/%Y")
    assert data["config"].get("report_date") == today


def test_template_kpis_no_debug_subtitle(tmp_path):
    """Nenhum KPI deve ter subtítulo de texto interno."""
    p = tmp_path / "test.xlsx"
    create_template(str(p))
    data, _ = read_excel(str(p))
    for kpi in data["kpis"]:
        sub = kpi.get("subtitulo") or ""
        assert sub != "mantido do modelo atual", "Texto interno de debug encontrado no subtítulo do KPI"


def test_template_pendencias_has_all_columns(tmp_path):
    """Cada pendência no template deve ter 13 colunas definidas."""
    p = tmp_path / "test.xlsx"
    create_template(str(p))
    wb = openpyxl.load_workbook(str(p))
    ws = wb["PENDENCIAS_CRITICAS"]
    headers = [c.value for c in ws[1] if c.value]
    expected = [
        "prioridade", "item", "responsaveis", "status", "nivel",
        "id_origem", "categoria", "score", "probabilidade", "impacto",
        "estrategia", "data_limite", "comentarios",
    ]
    assert headers == expected
    wb.close()


def test_template_rodape_simplified(tmp_path):
    """RODAPE do template não deve ter campos duplicados do CONFIG."""
    p = tmp_path / "test.xlsx"
    create_template(str(p))
    data, _ = read_excel(str(p))
    redundant = {"owner_relatorio", "nome_relatorio", "data_relatorio"}
    overlap = redundant.intersection(set(data["rodape"].keys()))
    assert not overlap, f"Campos redundantes encontrados no RODAPE: {overlap}"


def test_template_curva_s_max_100(tmp_path):
    """Nenhum ponto da Curva S deve ultrapassar 100%."""
    p = tmp_path / "test.xlsx"
    create_template(str(p))
    data, _ = read_excel(str(p))
    for pt in data["curva_s"]:
        assert (pt.get("planejado") or 0) <= 100, f"planejado > 100: {pt}"
        assert (pt.get("realizado") or 0) <= 100, f"realizado > 100: {pt}"


def test_missing_required_sheet_returns_empty_list(tmp_path):
    """Aba ausente deve retornar lista vazia sem erro de leitura."""
    p = tmp_path / "partial.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CONFIG"
    ws.append(["project_name", "Teste"])
    wb.save(str(p))
    data, error = read_excel(str(p))
    assert data is not None
    assert data["kpis"] == []
    assert data["fases"] == []


def test_cache_fallback_on_corrupt_file(tmp_path):
    """Após leitura bem-sucedida, cache deve ser retornado se Excel ficar ilegível."""
    p = tmp_path / "test.xlsx"
    create_template(str(p))
    # Primeira leitura — popula o cache
    data1, err1 = read_excel(str(p))
    assert err1 is None
    assert data1 is not None
    # Corrompe o arquivo (sobreescreve com bytes inválidos)
    p.write_bytes(b"INVALID CONTENT - NOT AN XLSX")
    # Segunda leitura — deve usar cache pois Excel está corrompido
    data2, err2 = read_excel(str(p))
    assert data2 is not None
    assert err2 is not None
    assert "cache" in err2.lower() or "cache válido" in err2.lower()


