"""Testes para backend/schema_validator.py"""
import pytest
import openpyxl
from backend.excel_reader import create_template
from backend.schema_validator import validate_schema


def test_valid_template_no_errors(tmp_path):
    p = tmp_path / "test.xlsx"
    create_template(str(p))
    errors = validate_schema(str(p))
    assert errors == [], f"Erros inesperados no template válido: {errors}"


def test_missing_required_sheet(tmp_path):
    p = tmp_path / "test.xlsx"
    create_template(str(p))
    wb = openpyxl.load_workbook(str(p))
    del wb["KPIS"]
    wb.save(str(p))
    wb.close()
    errors = validate_schema(str(p))
    assert any("KPIS" in e for e in errors)


def test_missing_required_column(tmp_path):
    p = tmp_path / "test.xlsx"
    create_template(str(p))
    wb = openpyxl.load_workbook(str(p))
    ws = wb["FASES"]
    for cell in ws[1]:
        if cell.value == "status":
            cell.value = "status_REMOVIDO"
    wb.save(str(p))
    wb.close()
    errors = validate_schema(str(p))
    assert any("status" in e.lower() for e in errors)


def test_config_fields_validated(tmp_path):
    """CONFIG deve ter validação dos campos chave-valor obrigatórios."""
    p = tmp_path / "test.xlsx"
    create_template(str(p))
    wb = openpyxl.load_workbook(str(p))
    ws = wb["CONFIG"]
    for row in ws.iter_rows(min_row=1, max_col=1):
        if row[0].value == "project_name":
            row[0].value = "project_name_REMOVIDO"
            break
    wb.save(str(p))
    wb.close()
    errors = validate_schema(str(p))
    assert any("project_name" in e for e in errors)


def test_nonexistent_file_returns_error():
    errors = validate_schema("arquivo_que_nao_existe.xlsx")
    assert len(errors) > 0
