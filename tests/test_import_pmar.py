"""Testes para backend/import_pmar_raid.py"""
import pytest
from pathlib import Path
from datetime import date
import openpyxl

from backend.excel_reader import create_template
from backend.import_pmar_raid import (
    clean, fmt_date, translate_status, map_priority, pluralize,
    _derive_current_phase,
)

PMAR_PATH = Path("docs/PMAR.45_RAID_Log.xlsm")


# ---------------------------------------------------------------------------
# Utilitários
# ---------------------------------------------------------------------------

def test_clean_none():
    assert clean(None) is None

def test_clean_error_strings():
    for err in ("#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"):
        assert clean(err) is None, f"clean({err!r}) deveria retornar None"

def test_clean_strips_whitespace():
    assert clean("  valor  ") == "valor"

def test_clean_empty_string_returns_none():
    assert clean("   ") is None

def test_fmt_date_date_object():
    from datetime import date
    d = date(2026, 5, 14)
    assert fmt_date(d) == "14/05/2026"

def test_fmt_date_datetime_object():
    from datetime import datetime
    dt = datetime(2026, 5, 14, 10, 30)
    assert fmt_date(dt) == "14/05/2026"

def test_fmt_date_none():
    assert fmt_date(None) is None

def test_fmt_date_string_passthrough():
    assert fmt_date("14/05/2026") == "14/05/2026"

def test_translate_status_known():
    assert translate_status("in progress") == "Em atenção"
    assert translate_status("resolved")    == "Concluído"
    assert translate_status("not started") == "Planejado"
    assert translate_status("cancelled")   == "Cancelado"

def test_translate_status_unknown_passthrough():
    assert translate_status("algum status novo") == "algum status novo"

def test_map_priority_critical():
    label, nivel = map_priority("critical")
    assert label == "P1"
    assert nivel == "danger"

def test_map_priority_high():
    label, nivel = map_priority("high")
    assert label == "P2"
    assert nivel == "warning"

def test_map_priority_unknown():
    label, nivel = map_priority("unknown")
    assert nivel == "gray"

def test_pluralize():
    assert pluralize(1, "risco", "riscos") == "risco"
    assert pluralize(0, "risco", "riscos") == "riscos"
    assert pluralize(2, "risco", "riscos") == "riscos"


# ---------------------------------------------------------------------------
# _derive_current_phase
# ---------------------------------------------------------------------------

def test_derive_current_phase_finds_active(tmp_path):
    p = tmp_path / "test.xlsx"
    create_template(str(p))
    wb = openpyxl.load_workbook(str(p))
    # Template já tem "Explorer" como "Em andamento"
    phase = _derive_current_phase(wb)
    wb.close()
    assert phase == "Explorer"

def test_derive_current_phase_no_fases(tmp_path):
    p = tmp_path / "test.xlsx"
    create_template(str(p))
    wb = openpyxl.load_workbook(str(p))
    del wb["FASES"]
    phase = _derive_current_phase(wb)
    wb.close()
    assert phase is None

def test_derive_current_phase_all_planejado(tmp_path):
    p = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FASES"
    ws.append(["ordem", "nome", "status", "data_alvo", "destaque"])
    ws.append([1, "Fase A", "Planejado", "01/06/2026", False])
    ws.append([2, "Fase B", "Planejado", "01/07/2026", False])
    wb.save(str(p))
    wb2 = openpyxl.load_workbook(str(p))
    phase = _derive_current_phase(wb2)
    wb2.close()
    assert phase is None


# ---------------------------------------------------------------------------
# Integração com PMAR real (pula se o arquivo não existir)
# ---------------------------------------------------------------------------

@pytest.mark.skipif(not PMAR_PATH.exists(), reason="PMAR real não disponível")
def test_read_pmar_returns_expected_keys():
    from backend.import_pmar_raid import read_pmar
    result = read_pmar(PMAR_PATH)
    for key in ("config", "team", "risks", "actions", "issues", "decisions"):
        assert key in result, f"Chave '{key}' ausente no resultado de read_pmar"

@pytest.mark.skipif(not PMAR_PATH.exists(), reason="PMAR real não disponível")
def test_import_pmar_idempotente(tmp_path):
    """Rodar a importação duas vezes não deve duplicar dados."""
    from backend.import_pmar_raid import import_pmar
    status = tmp_path / "status.xlsx"
    create_template(str(status))
    result1 = import_pmar(PMAR_PATH, status)
    result2 = import_pmar(PMAR_PATH, status)
    assert result1["riscos_total"]   == result2["riscos_total"]
    assert result1["riscos_abertos"] == result2["riscos_abertos"]

@pytest.mark.skipif(not PMAR_PATH.exists(), reason="PMAR real não disponível")
def test_import_pmar_no_debug_text(tmp_path):
    """Após importar, nenhum KPI deve ter o subtítulo de debug."""
    from backend.import_pmar_raid import import_pmar
    from backend.excel_reader import read_excel
    import backend.excel_reader as em
    em._cache = None

    status = tmp_path / "status.xlsx"
    create_template(str(status))
    import_pmar(PMAR_PATH, status)
    data, _ = read_excel(str(status))
    for kpi in data["kpis"]:
        sub = kpi.get("subtitulo") or ""
        assert sub != "mantido do modelo atual"

@pytest.mark.skipif(not PMAR_PATH.exists(), reason="PMAR real não disponível")
def test_import_pmar_creates_backup(tmp_path):
    """Importação deve criar arquivo de backup."""
    from backend.import_pmar_raid import import_pmar
    status = tmp_path / "status.xlsx"
    create_template(str(status))
    import_pmar(PMAR_PATH, status)
    backup = tmp_path / "status_backup.xlsx"
    assert backup.exists(), "Arquivo de backup não foi criado"

@pytest.mark.skipif(not PMAR_PATH.exists(), reason="PMAR real não disponível")
def test_import_pmar_project_subtitle_not_sponsor(tmp_path):
    """project_subtitle não deve receber o valor do Sponsor."""
    from backend.import_pmar_raid import import_pmar, read_pmar
    from backend.excel_reader import read_excel
    import backend.excel_reader as em
    em._cache = None

    status = tmp_path / "status.xlsx"
    create_template(str(status))
    pmar_data = read_pmar(PMAR_PATH)
    sponsor = pmar_data["config"].get("Sponsor", "")

    import_pmar(PMAR_PATH, status)
    em._cache = None
    data, _ = read_excel(str(status))
    subtitle = data["config"].get("project_subtitle", "")
    assert subtitle != sponsor, (
        f"project_subtitle ({subtitle!r}) não deve ser igual ao Sponsor ({sponsor!r})"
    )
